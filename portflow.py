import requests
import csv
import time
import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path


def load_env_file(env_path: Path | None = None) -> None:
    env_file = env_path or (Path(__file__).resolve().parent / ".env")
    if not env_file.exists():
        return

    raw_lines = env_file.read_text(encoding="utf-8").splitlines()
    i = 0
    while i < len(raw_lines):
        line = raw_lines[i].strip()
        i += 1

        if not line or line.startswith("#"):
            continue
        if line.startswith("export "):
            line = line[7:].strip()
        if "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if not key:
            continue

        # Support multi-line values that start with [ or {
        if value and value[0] in ("[", "{"):
            depth = value.count("[") + value.count("{") - value.count("]") - value.count("}")
            while depth > 0 and i < len(raw_lines):
                chunk = raw_lines[i].strip()
                i += 1
                value += chunk
                depth += chunk.count("[") + chunk.count("{") - chunk.count("]") - chunk.count("}")
        elif len(value) >= 2 and value[0] == value[-1] and value[0] in {'"', "'"}:
            value = value[1:-1]

        os.environ[key] = value


def parse_coach_students(raw_json: str):
    if not raw_json.strip():
        return []

    import re as _re
    raw_json = _re.sub(r",\s*([\]\}])", r"\1", raw_json)  # strip trailing commas
    try:
        data = json.loads(raw_json)
    except json.JSONDecodeError:
        print("Waarschuwing: PORTFLOW_COACH_STUDENTS_JSON bevat ongeldige JSON. Gebruik lege lijst.")
        return []

    if not isinstance(data, list):
        print("Waarschuwing: PORTFLOW_COACH_STUDENTS_JSON moet een JSON lijst zijn. Gebruik lege lijst.")
        return []

    students = []
    for entry in data:
        if not isinstance(entry, dict):
            continue

        # Separator entry: {"separator": true}
        if entry.get("separator") is True:
            students.append((SEPARATOR_SENTINEL, None, "", "", "", ""))
            continue

        name = str(entry.get("name") or "").strip()
        if not name:
            continue

        start_date_raw = str(entry.get("start_date") or "").strip()
        start_date = None
        if start_date_raw:
            try:
                start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
            except ValueError:
                print(
                    "Waarschuwing: ongeldige startdatum '",
                    start_date_raw,
                    "' voor student ",
                    name,
                    ". Verwacht YYYY-MM-DD.",
                    sep="",
                )

        semester = str(entry.get("semester") or "").strip()
        tribe = str(entry.get("tribe") or "").strip()
        gilde = str(entry.get("gilde") or "").strip()
        coach = str(entry.get("coach") or "").strip()

        students.append((name, start_date, semester, tribe, gilde, coach))

    return students


print(r"  ____            _    __ _                ")
print(r" |  _ \ ___  _ __| |_ / _| | _____      __ ")
print(r" | |_) / _ \| '__| __| |_| |/ _ \ \ /\ / / ")
print(r" |  __/ (_) | |  | |_|  _| | (_) \ V  V /  ")
print(r" |_|   \___/|_|   \__|_| |_|\___/ \_/\_/   ")
print(r"      dashboard  —  peter.snoek@hu.nl      ")
print()

# Controleer vereiste packages
_REQUIRED_PACKAGES = [
    ("requests",  "requests",  "pip install requests"),
    ("openpyxl",  "openpyxl",  "pip install openpyxl"),
]
_missing = []
for _pkg_import, _pkg_name, _pkg_install in _REQUIRED_PACKAGES:
    try:
        __import__(_pkg_import)
    except ImportError:
        _missing.append((_pkg_name, _pkg_install))
if _missing:
    print("Ontbrekende packages gevonden. Installeer ze met:")
    for _name, _cmd in _missing:
        print(f"  {_cmd}    # {_name}")
    print()
    if any(n == "requests" for n, _ in _missing):
        print("Het script kan niet verder zonder 'requests'. Gestopt.")
        sys.exit(1)
    else:
        print("Tip: zonder 'openpyxl' is het inlezen van Excel-bestanden niet beschikbaar.")
        print()

load_env_file()

BASE_URL = "https://portfolio.drieam.app/api/v1"
PER_PAGE = 200
BEARER_TOKEN = os.getenv("PORTFLOW_BEARER_TOKEN", "").strip()

SECTION_ID = "72086"


def _load_semester_dates() -> tuple:
    """Lees semester datums uit .env; vraag interactief als ze ontbreken."""
    env_file = Path(__file__).resolve().parent / ".env"

    def _read_date(key: str, prompt: str):
        val = os.getenv(key, "").strip()
        if val:
            try:
                return datetime.strptime(val, "%Y-%m-%d").date()
            except ValueError:
                print(f"Waarschuwing: {key} heeft ongeldige datum '{val}'. Verwacht YYYY-MM-DD.")
        # Ontbreekt of ongeldig — vraag interactief
        while True:
            answer = input(f"{prompt} (YYYY-MM-DD): ").strip()
            try:
                date_obj = datetime.strptime(answer, "%Y-%m-%d").date()
                # Schrijf terug naar .env
                if env_file.exists():
                    text = env_file.read_text(encoding="utf-8")
                    import re as _re2
                    pattern = rf"^{key}=.*$"
                    replacement = f"{key}={answer}"
                    if _re2.search(pattern, text, flags=_re2.MULTILINE):
                        text = _re2.sub(pattern, replacement, text, flags=_re2.MULTILINE)
                    else:
                        text = text.rstrip("\n") + f"\n{replacement}\n"
                    env_file.write_text(text, encoding="utf-8")
                else:
                    env_file.write_text(f"{key}={answer}\n", encoding="utf-8")
                os.environ[key] = answer
                return date_obj
            except ValueError:
                print("Ongeldige datum. Voer in als YYYY-MM-DD, bijv. 2026-02-12.")

    start = _read_date("PORTFLOW_SEMESTER_START", "Op welke dag begint het semester")
    end = _read_date("PORTFLOW_SEMESTER_END", "Wat is de laatste dag van het semester")
    return start, end


CURRENT_SEMESTER_START, CURRENT_SEMESTER_END = _load_semester_dates()
print(f"Semester periode: {CURRENT_SEMESTER_START.strftime('%d-%m-%Y')} t/m {CURRENT_SEMESTER_END.strftime('%d-%m-%Y')}")

GOALS = {
    "Overzicht creëren",
    "Kritisch oordelen",
    "Juiste kennis ontwikkelen",
    "Kwalitatief Product Maken",
    "Plannen",
    "Boodschap Delen",
    "Samenwerken",
    "Flexibel opstellen",
    "Pro-actief handelen",
    "Reflecteren"
}

GOAL_COLUMNS = [
    ("Overzicht creëren",           "OC"),
    ("Kritisch oordelen",           "KO"),
    ("Juiste kennis ontwikkelen",   "JKO"),
    ("Kwalitatief Product Maken",   "KPM"),
    ("Plannen",                     "PL"),
    ("Boodschap Delen",             "BD"),
    ("Samenwerken",                 "SW"),
    ("Flexibel opstellen",          "FO"),
    ("Pro-actief handelen",         "PH"),
    ("Reflecteren",                 "RE"),
]

# Sentinel waarde in .env die een horizontale scheidingslijn in de tabel veroorzaakt
SEPARATOR_SENTINEL = "---"

# JSON lijst met studentconfig uit .env.
# Voorbeeld item: {"name":"Student Naam","start_date":"2026-04-18"}
CURRENT_COACH_STUDENTS = parse_coach_students(
    os.getenv("PORTFLOW_COACH_STUDENTS_JSON", "").strip()
)

# Afgeleide lookups voor snel gebruik
COACH_STUDENT_NAMES       = {name for name, *_ in CURRENT_COACH_STUDENTS}
COACH_STUDENT_START_DATES = {name: start for name, start, *_ in CURRENT_COACH_STUDENTS if start is not None}
COACH_STUDENT_SEMESTER    = {name: sem   for name, _, sem, *_      in CURRENT_COACH_STUDENTS if sem}
COACH_STUDENT_TRIBE       = {name: tribe for name, _, _, tribe, *_ in CURRENT_COACH_STUDENTS if tribe}
COACH_STUDENT_GILDE       = {name: gilde for name, _, _, _, gilde, *_ in CURRENT_COACH_STUDENTS if gilde}
COACH_STUDENT_COACH       = {name: coach for name, _, _, _, _, coach   in CURRENT_COACH_STUDENTS if coach}

# Tribe-brede lijst (eigen studenten + collega's) uit .env
CURRENT_TRIBE_STUDENTS = parse_coach_students(
    os.getenv("PORTFLOW_TRIBE_STUDENTS_JSON", "").strip()
)
TRIBE_STUDENT_NAMES       = {name for name, *_ in CURRENT_TRIBE_STUDENTS}
TRIBE_STUDENT_START_DATES = {name: start for name, start, *_ in CURRENT_TRIBE_STUDENTS if start is not None}
TRIBE_STUDENT_SEMESTER    = {name: sem   for name, _, sem, *_   in CURRENT_TRIBE_STUDENTS if sem}
TRIBE_STUDENT_TRIBE       = {name: tribe for name, _, _, tribe, *_ in CURRENT_TRIBE_STUDENTS if tribe}
TRIBE_STUDENT_GILDE       = {name: gilde for name, _, _, _, gilde, *_ in CURRENT_TRIBE_STUDENTS if gilde}
TRIBE_STUDENT_COACH       = {name: coach for name, _, _, _, _, coach  in CURRENT_TRIBE_STUDENTS if coach}

# Gilde-brede lijst uit .env
CURRENT_GILDE_STUDENTS = parse_coach_students(
    os.getenv("PORTFLOW_GILDE_STUDENTS_JSON", "").strip()
)
GILDE_STUDENT_NAMES       = {name for name, *_ in CURRENT_GILDE_STUDENTS}
GILDE_STUDENT_START_DATES = {name: start for name, start, *_ in CURRENT_GILDE_STUDENTS if start is not None}
GILDE_STUDENT_SEMESTER    = {name: sem   for name, _, sem, *_   in CURRENT_GILDE_STUDENTS if sem}
GILDE_STUDENT_TRIBE       = {name: tribe for name, _, _, tribe, *_ in CURRENT_GILDE_STUDENTS if tribe}
GILDE_STUDENT_GILDE       = {name: gilde for name, _, _, _, gilde, *_ in CURRENT_GILDE_STUDENTS if gilde}
GILDE_STUDENT_COACH       = {name: coach for name, _, _, _, _, coach  in CURRENT_GILDE_STUDENTS if coach}

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--dump-schema",
        action="store_true",
        help="Schrijf een overzicht van alle gevonden API-velden naar schema_inventory.txt",
    )
    parser.add_argument(
        "--debug-api",
        action="store_true",
        help="Log alle API-calls en responses naar api_debug_log.jsonl",
    )
    parser.add_argument(
        "--debug-pending",
        action="store_true",
        help="Log evaluatie-beslissingen (waarom wel/geen ?) naar pending_debug.json",
    )
    parser.add_argument(
        "--anoniem",
        action="store_true",
        help="Vervang alle studentnamen door ******* in de uitvoer",
    )
    parser.add_argument(
        "--vraagtekens",
        choices=["1"],
        default=None,
        help="Toon openstaande (?) evaluaties: 1=tonen",
    )
    parser.add_argument(
        "--aantal",
        choices=["1", "meer"],
        default=None,
        help="Hoeveel studenten: 1=één student, meer=meerdere studenten",
    )
    parser.add_argument(
        "--lijst",
        choices=["alles", "coach", "tribe", "gilde"],
        default=None,
        help="Welke studentenlijst: alles=gedeeld, coach=mijn coach studenten, tribe=tribe, gilde=gilde",
    )
    parser.add_argument(
        "--student",
        default=None,
        metavar="NUMMER",
        help="Kies direct de Nde student uit de lijst (volgnummer, alleen bij --aantal 1)",
    )
    parser.add_argument(
        "--admin",
        action="store_true",
        help="Toon beheeropties (Excel-import, studenten toevoegen aan .env)",
    )
    return parser.parse_args()


ARGS = parse_args()
DUMP_SCHEMA = ARGS.dump_schema
DEBUG_API = ARGS.debug_api
DEBUG_PENDING = ARGS.debug_pending

API_DEBUG_FILE = "api_debug_log.jsonl"
PENDING_DEBUG_FILE = "pending_debug.json"

PENDING_DEBUG_EVENTS = []


def init_debug_files():
    if DEBUG_API:
        with open(API_DEBUG_FILE, "w", encoding="utf-8") as f:
            f.write("")
        print(f"API debug logging enabled: {API_DEBUG_FILE}")

    if DEBUG_PENDING:
        with open(PENDING_DEBUG_FILE, "w", encoding="utf-8") as f:
            f.write("[]\n")
        print(f"Pending debug logging enabled: {PENDING_DEBUG_FILE}")


def log_api_debug(url, params, status_code=None, body=None, error=None):
    if not DEBUG_API:
        return

    payload = {
        "timestamp": datetime.now().isoformat(),
        "url": url,
        "params": params,
        "status_code": status_code,
        "error": error,
        "response_body": body,
    }

    with open(API_DEBUG_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(payload, ensure_ascii=False) + "\n")


def log_pending_debug_event(event):
    if not DEBUG_PENDING:
        return
    PENDING_DEBUG_EVENTS.append(event)


def maybe_write_pending_debug_report():
    if not DEBUG_PENDING:
        return
    with open(PENDING_DEBUG_FILE, "w", encoding="utf-8") as f:
        json.dump(PENDING_DEBUG_EVENTS, f, ensure_ascii=False, indent=2)
        f.write("\n")


class SchemaInventory:
    def __init__(self):
        self.path_types = {}
        self.path_samples = {}

    def _type_name(self, value):
        if value is None:
            return "null"
        if isinstance(value, bool):
            return "bool"
        if isinstance(value, int):
            return "int"
        if isinstance(value, float):
            return "float"
        if isinstance(value, str):
            return "str"
        if isinstance(value, list):
            return "list"
        if isinstance(value, dict):
            return "dict"
        return type(value).__name__

    def _add_observation(self, path, value):
        t = self._type_name(value)
        self.path_types.setdefault(path, set()).add(t)

        if t in {"str", "int", "float", "bool", "null"}:
            sample = repr(value)
            samples = self.path_samples.setdefault(path, set())
            if len(samples) < 3:
                samples.add(sample)

    def observe(self, value, path):
        self._add_observation(path, value)

        if isinstance(value, dict):
            for k, v in value.items():
                self.observe(v, f"{path}.{k}")
            return

        if isinstance(value, list):
            for entry in value:
                self.observe(entry, f"{path}[]")

    def write_report(self, file_path="schema_inventory.txt"):
        lines = []
        lines.append("# API Schema Inventory (observed from live responses)")
        lines.append("")
        lines.append(
            "# Tip: run with option 'All students' for a more complete overview."
        )
        lines.append("")

        for path in sorted(self.path_types.keys()):
            types = ", ".join(sorted(self.path_types[path]))
            samples = sorted(self.path_samples.get(path, set()))
            if samples:
                lines.append(f"{path} | types={types} | samples={'; '.join(samples)}")
            else:
                lines.append(f"{path} | types={types}")

        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines) + "\n")


SCHEMA = SchemaInventory() if DUMP_SCHEMA else None
init_debug_files()


def observe_schema(value, path):
    if SCHEMA is not None:
        SCHEMA.observe(value, path)


def maybe_write_schema_report():
    if SCHEMA is None:
        return
    SCHEMA.write_report("schema_inventory.txt")
    print("Schema inventory exported to schema_inventory.txt")



def _save_token_to_env(token: str) -> None:
    """Schrijft de bearer token naar .env (vervangt een eventuele bestaande waarde)."""
    env_path = Path(__file__).resolve().parent / ".env"
    if env_path.exists():
        content = env_path.read_text(encoding="utf-8")
        if re.search(r"^PORTFLOW_BEARER_TOKEN=", content, flags=re.MULTILINE):
            content = re.sub(
                r"^PORTFLOW_BEARER_TOKEN=.*$",
                f"PORTFLOW_BEARER_TOKEN={token}",
                content,
                flags=re.MULTILINE,
            )
        else:
            content = f"PORTFLOW_BEARER_TOKEN={token}\n" + content
    else:
        content = f"PORTFLOW_BEARER_TOKEN={token}\n"
    env_path.write_text(content, encoding="utf-8")
    print("  Token opgeslagen in .env")


def _jwt_is_valid(token: str) -> bool:
    """Controleer of een JWT-token nog niet verlopen is op basis van de 'exp' claim."""
    try:
        import base64 as _b64
        payload_b64 = token.split(".")[1]
        padding = (4 - len(payload_b64) % 4) % 4
        payload = json.loads(_b64.urlsafe_b64decode(payload_b64 + "=" * padding))
        exp = payload.get("exp")
        if exp is None:
            return True  # geen exp claim → aannemen dat het geldig is
        return time.time() < exp
    except Exception:
        return False  # bij twijfel → als verlopen beschouwen


def _jwt_user_name(token: str) -> str | None:
    """Haal de gebruikersnaam op uit het JWT-payload (name / sub / email)."""
    try:
        import base64 as _b64
        payload_b64 = token.split(".")[1]
        padding = (4 - len(payload_b64) % 4) % 4
        payload = json.loads(_b64.urlsafe_b64decode(payload_b64 + "=" * padding))
        for key in ("name", "full_name", "fullname", "display_name", "email"):
            val = payload.get(key)
            if val and isinstance(val, str) and val.strip():
                return val.strip()
        # Zoek recursief in geneste objecten
        for val in payload.values():
            if isinstance(val, dict):
                for key in ("name", "full_name", "email"):
                    v = val.get(key)
                    if v and isinstance(v, str) and v.strip():
                        return v.strip()
    except Exception:
        pass
    return None


OWN_PORTFOLIO_ID = os.getenv("PORTFLOW_OWN_PORTFOLIO_ID", "").strip()


def _get_own_user_name(token: str) -> str | None:
    """Haal de naam van de ingelogde gebruiker op via de eigen portfolio API."""
    if not OWN_PORTFOLIO_ID:
        return _jwt_user_name(token)
    try:
        headers = {"accept": "*/*", "authorization": f"Bearer {token}", "user-agent": "Mozilla/5.0"}
        response = requests.get(
            f"{BASE_URL}/portfolios/{OWN_PORTFOLIO_ID}",
            headers=headers,
            timeout=5,
        )
        if response.status_code == 200:
            data = response.json()
            user = data.get("user") or {}
            name = user.get("name") or user.get("full_name") or user.get("email")
            if name and isinstance(name, str) and name.strip():
                return name.strip()
    except Exception:
        pass
    return _jwt_user_name(token)


def get_bearer_token(force_prompt: bool = False):
    token_from_var = (globals().get("BEARER_TOKEN") or "").strip()

    # Stap 1: gebruik token uit .env als het nog geldig is
    if token_from_var and not force_prompt:
        if _jwt_is_valid(token_from_var):
            print("Authentication bearer token opgehaald uit .env")
            return token_from_var
        print("Authentication bearer token uit .env is verouderd.")
    elif token_from_var and force_prompt:
        print("Authentication bearer token uit .env is verouderd.")

    print()
    print("Haal een nieuw token op via browser inspector → Network → filter op 'dashboard'")
    print()
    token = input("Plak bearer token: ").strip()
    if not token:
        raise SystemExit("Geen token opgegeven, script gestopt.")

    save = input("Token opslaan in .env? [J/n] ").strip().lower()
    if save in ("", "j", "y", "yes", "ja"):
        _save_token_to_env(token)
        globals()["BEARER_TOKEN"] = token
    return token


def get_section_id(force_prompt: bool = False) -> str:
    section_from_var = (globals().get("SECTION_ID") or "").strip()
    if section_from_var and not force_prompt:
        print(f"Using SECTION_ID from variable: {section_from_var}")
        return section_from_var
    if not section_from_var:
        print("No SECTION_ID set; prompting for section_id...")
    return input("Enter section_id: ").strip()


def anonymize_name(full_name: str) -> str:
    if not isinstance(full_name, str) or not full_name.strip():
        return full_name
    return " ".join(p[0] + "*" * (len(p) - 1) for p in full_name.split())


def name_to_initials(full_name: str) -> str:
    if not isinstance(full_name, str):
        return ""
    # Drop common Dutch tussenvoegsels.
    skip = {"van", "de", "der", "den", "ten", "ter", "te", "'t", "v", "d"}
    parts = [p for p in full_name.replace("-", " ").split() if p]
    letters = []
    for part in parts:
        lowered = part.lower()
        if lowered in skip:
            continue
        # Keep only alphabetic start.
        first = next((ch for ch in part if ch.isalpha()), "")
        if first:
            letters.append(first.upper())
    return "".join(letters)


def student_order_and_width(students: dict, preferred_order: list[str] | None = None) -> tuple[list[str], int]:
    if preferred_order:
        ordered_names = [n for n in preferred_order if n in students]
        # append any names not in preferred_order at the end
        ordered_names += sorted(n for n in students if n not in set(ordered_names))
    else:
        ordered_names = sorted(students.keys())
    width = max(2, len(str(len(ordered_names))))
    return ordered_names, width


def resolve_student_selection(selection: str, ordered_names: list[str], students: dict) -> str | None:
    s = (selection or "").strip()
    if not s:
        return None

    if s.isdigit():
        idx = int(s)
        if 1 <= idx <= len(ordered_names):
            return ordered_names[idx - 1]
        return None

    if s in students:
        return s

    return None


def choose_semester_scope() -> str:
    return "current"


# ------------------------
# Excel import helpers
# ------------------------

def _find_excel_with_student_columns() -> "list[tuple[Path, str]]":
    """Geeft alle .xlsx-bestanden in de scriptmap die de vereiste kolomkoppen bevatten."""
    required = {"Nummer", "Naam", "Email", "Semester", "Project", "Coach", "Gilde voorkeur"}
    script_dir = Path(__file__).resolve().parent
    matches = []
    for path in sorted(script_dir.glob("*.xlsx")):
        try:
            import openpyxl as _xl
            wb = _xl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            headers = {c.value for c in next(ws.iter_rows(min_row=1, max_row=1)) if c.value}
            wb.close()
            if required.issubset(headers):
                matches.append((path, path.name))
        except Exception:
            continue
    return matches


def _extract_semester_number(sem_str: str) -> str:
    m = re.search(r"\d+", str(sem_str or ""))
    return m.group(0) if m else ""


def _extract_tribe_from_project(project_str: str) -> str:
    """Geeft projectnummer + opdrachtgever terug, bijv. '(08) Commerciele economie'."""
    s = re.sub(r"\s*\([^)]*\)\s*$", "", str(project_str or "")).strip()
    dash = s.find(" - ")
    return s[:dash].strip() if dash != -1 else s


def _group_by_tribe(entries: list) -> list:
    """Sorteert entries op tribe (dan naam) en voegt separators in tussen tribes."""
    sorted_entries = sorted(entries, key=lambda e: (e.get("tribe") or "", e.get("name") or ""))
    result = []
    prev_tribe = None
    for e in sorted_entries:
        tribe = e.get("tribe") or ""
        if prev_tribe is not None and tribe != prev_tribe:
            result.append({"separator": True})
        result.append(e)
        prev_tribe = tribe
    return result


def _write_env_key(key: str, value: str) -> None:
    """Vervangt of voegt een (mogelijk meerregelige) sleutel toe in .env."""
    env_path = Path(__file__).resolve().parent / ".env"
    new_entry = f"{key}={value}\n"
    if not env_path.exists():
        env_path.write_text(new_entry, encoding="utf-8")
        return
    content = env_path.read_text(encoding="utf-8")
    m = re.search(rf"^{re.escape(key)}=", content, flags=re.MULTILINE)
    if m is None:
        content = content.rstrip("\n") + "\n" + new_entry
    else:
        search_from = m.end()
        nxt = re.search(r"\n[A-Z][A-Z0-9_]+=", content[search_from:])
        end = search_from + nxt.start() + 1 if nxt else len(content)
        content = content[: m.start()] + new_entry + content[end:]
    env_path.write_text(content, encoding="utf-8")


def _reload_coach_students() -> None:
    """Herlaadt PORTFLOW_COACH_STUDENTS_JSON vanuit os.environ in de globale lookups."""
    global CURRENT_COACH_STUDENTS, COACH_STUDENT_NAMES, COACH_STUDENT_START_DATES
    global COACH_STUDENT_SEMESTER, COACH_STUDENT_TRIBE, COACH_STUDENT_GILDE, COACH_STUDENT_COACH
    CURRENT_COACH_STUDENTS = parse_coach_students(
        os.getenv("PORTFLOW_COACH_STUDENTS_JSON", "").strip()
    )
    COACH_STUDENT_NAMES = {name for name, *_ in CURRENT_COACH_STUDENTS}
    COACH_STUDENT_START_DATES = {
        name: start for name, start, *_ in CURRENT_COACH_STUDENTS if start is not None
    }
    COACH_STUDENT_SEMESTER = {name: sem   for name, _, sem, *_      in CURRENT_COACH_STUDENTS if sem}
    COACH_STUDENT_TRIBE    = {name: tribe for name, _, _, tribe, *_ in CURRENT_COACH_STUDENTS if tribe}
    COACH_STUDENT_GILDE    = {name: gilde for name, _, _, _, gilde, *_ in CURRENT_COACH_STUDENTS if gilde}
    COACH_STUDENT_COACH    = {name: coach for name, _, _, _, _, coach   in CURRENT_COACH_STUDENTS if coach}


def _reload_tribe_students() -> None:
    """Herlaadt PORTFLOW_TRIBE_STUDENTS_JSON vanuit os.environ in de globale lookups."""
    global CURRENT_TRIBE_STUDENTS, TRIBE_STUDENT_NAMES, TRIBE_STUDENT_START_DATES
    global TRIBE_STUDENT_SEMESTER, TRIBE_STUDENT_TRIBE, TRIBE_STUDENT_GILDE, TRIBE_STUDENT_COACH
    CURRENT_TRIBE_STUDENTS = parse_coach_students(
        os.getenv("PORTFLOW_TRIBE_STUDENTS_JSON", "").strip()
    )
    TRIBE_STUDENT_NAMES       = {name for name, *_ in CURRENT_TRIBE_STUDENTS}
    TRIBE_STUDENT_START_DATES = {name: start for name, start, *_ in CURRENT_TRIBE_STUDENTS if start is not None}
    TRIBE_STUDENT_SEMESTER    = {name: sem   for name, _, sem, *_      in CURRENT_TRIBE_STUDENTS if sem}
    TRIBE_STUDENT_TRIBE       = {name: tribe for name, _, _, tribe, *_ in CURRENT_TRIBE_STUDENTS if tribe}
    TRIBE_STUDENT_GILDE       = {name: gilde for name, _, _, _, gilde, *_ in CURRENT_TRIBE_STUDENTS if gilde}
    TRIBE_STUDENT_COACH       = {name: coach for name, _, _, _, _, coach   in CURRENT_TRIBE_STUDENTS if coach}


def _reload_gilde_students() -> None:
    """Herlaadt PORTFLOW_GILDE_STUDENTS_JSON vanuit os.environ in de globale lookups."""
    global CURRENT_GILDE_STUDENTS, GILDE_STUDENT_NAMES, GILDE_STUDENT_START_DATES
    global GILDE_STUDENT_SEMESTER, GILDE_STUDENT_TRIBE, GILDE_STUDENT_GILDE, GILDE_STUDENT_COACH
    CURRENT_GILDE_STUDENTS = parse_coach_students(
        os.getenv("PORTFLOW_GILDE_STUDENTS_JSON", "").strip()
    )
    GILDE_STUDENT_NAMES       = {name for name, *_ in CURRENT_GILDE_STUDENTS}
    GILDE_STUDENT_START_DATES = {name: start for name, start, *_ in CURRENT_GILDE_STUDENTS if start is not None}
    GILDE_STUDENT_SEMESTER    = {name: sem   for name, _, sem, *_      in CURRENT_GILDE_STUDENTS if sem}
    GILDE_STUDENT_TRIBE       = {name: tribe for name, _, _, tribe, *_ in CURRENT_GILDE_STUDENTS if tribe}
    GILDE_STUDENT_GILDE       = {name: gilde for name, _, _, _, gilde, *_ in CURRENT_GILDE_STUDENTS if gilde}
    GILDE_STUDENT_COACH       = {name: coach for name, _, _, _, _, coach   in CURRENT_GILDE_STUDENTS if coach}


def _migrate_env_add_coach_field() -> None:
    """Voegt het veld 'coach' toe aan bestaande entries in alle drie groepen (éénmalig)."""
    _keys = [
        ("PORTFLOW_COACH_STUDENTS_JSON", _reload_coach_students),
        ("PORTFLOW_TRIBE_STUDENTS_JSON", _reload_tribe_students),
        ("PORTFLOW_GILDE_STUDENTS_JSON", _reload_gilde_students),
    ]
    for env_key, reload_fn in _keys:
        raw = os.getenv(env_key, "").strip()
        if not raw:
            continue
        try:
            lst: list = json.loads(re.sub(r",\s*([\]\}])", r"\1", raw))
        except Exception:
            continue
        if not isinstance(lst, list):
            continue
        changed = False
        for entry in lst:
            if isinstance(entry, dict) and not entry.get("separator") and "coach" not in entry:
                entry["coach"] = ""
                changed = True
        if changed:
            parts = ["["]
            for i, e in enumerate(lst):
                comma = "," if i < len(lst) - 1 else ""
                parts.append(f"    {json.dumps(e, ensure_ascii=False)}{comma}")
            parts.append("    ]")
            json_str = "\n".join(parts)
            _write_env_key(env_key, json_str)
            os.environ[env_key] = json_str
            reload_fn()


def _pick_env_group_and_add(student_name: str) -> bool:
    """
    Toont een keuzemenu voor coach/tribe/gilde en voegt student_name toe
    aan de gekozen groep in .env. Geeft True terug bij succes.
    """
    _HL = "\033[48;2;100;160;220m\033[1m"
    _RS = "\033[0m"

    _groups = [
        ("PORTFLOW_COACH_STUDENTS_JSON", "coach studenten", _reload_coach_students),
        ("PORTFLOW_TRIBE_STUDENTS_JSON", "tribe studenten", _reload_tribe_students),
        ("PORTFLOW_GILDE_STUDENTS_JSON", "gilde studenten", _reload_gilde_students),
    ]
    print(f"\nAan welke groep wil je '{student_name}' toevoegen?")
    for i, (key, label, _) in enumerate(_groups, 1):
        print(f"  {i}) {label}  ({key})")
    print()

    raw = input("Keuze (1/2/3): ").strip()
    if not raw.isdigit() or not (1 <= int(raw) <= len(_groups)):
        print("Ongeldige keuze.")
        return False
    chosen_idx = int(raw)
    env_key, _, reload_fn = _groups[chosen_idx - 1]

    # Herdruk met highlight
    print(f"\033[{len(_groups) + 3}A\033[J", end="")
    print(f"\nAan welke groep wil je '{student_name}' toevoegen?")
    for i, (key, label, _) in enumerate(_groups, 1):
        if i == chosen_idx:
            print(f"  {_HL}{i}) {label}  ({key}){_RS}")
        else:
            print(f"  {i}) {label}  ({key})")
    print()
    print(f"Keuze (1/2/3): {raw}")

    existing_raw = os.getenv(env_key, "").strip()
    try:
        existing_list: list = json.loads(re.sub(r",\s*([\]\}])", r"\1", existing_raw)) if existing_raw else []
        if not isinstance(existing_list, list):
            existing_list = []
    except Exception:
        existing_list = []

    existing_names = {
        (e.get("name") or "").lower()
        for e in existing_list
        if isinstance(e, dict) and not e.get("separator")
    }
    if student_name.lower() in existing_names:
        print(f"  '{student_name}' staat al in {env_key}.")
        return False

    existing_list.append({"name": student_name, "start_date": None, "semester": "", "tribe": "", "gilde": "", "coach": ""})
    existing_list = _group_by_tribe([e for e in existing_list if not e.get("separator")])

    parts = ["["]
    for i, e in enumerate(existing_list):
        comma = "," if i < len(existing_list) - 1 else ""
        parts.append(f"    {json.dumps(e, ensure_ascii=False)}{comma}")
    parts.append("    ]")
    json_str = "\n".join(parts)

    _write_env_key(env_key, json_str)
    os.environ[env_key] = json_str
    reload_fn()
    print(f"  '{student_name}' toegevoegd aan {env_key}.")
    return True


def _pick_student_from_list(names: list, header: str, token: str) -> "tuple[str | None, str]":
    """
    Toont een genummerde lijst van namen en laat de gebruiker er een kiezen.
    Geeft (gekozen_naam, token) terug; naam is None bij annulering.
    """
    _HL = "\033[48;2;100;160;220m\033[1m"
    _RS = "\033[0m"

    number_width = max(2, len(str(len(names))))
    print(f"\n{header} ({len(names)} gevonden):")
    for i, name in enumerate(names, 1):
        print(f"  {i:0{number_width}d}) {name}")
    print()

    raw = input("Kies een student (volgnummer): ").strip()
    if not raw.isdigit() or not (1 <= int(raw) <= len(names)):
        print("Ongeldige keuze.")
        return None, token
    chosen_idx = int(raw)
    student_name = names[chosen_idx - 1]

    # Herdruk met highlight
    print(f"\033[{len(names) + 3}A\033[J", end="")
    print(f"\n{header} ({len(names)} gevonden):")
    for i, name in enumerate(names, 1):
        if i == chosen_idx:
            print(f"  {_HL}{i:0{number_width}d}) {name}{_RS}")
        else:
            print(f"  {i:0{number_width}d}) {name}")
    print()
    print(f"Kies een student (volgnummer): {raw}")

    return student_name, token


def _add_shared_student_to_env(token: str) -> tuple:
    """
    Haalt alle gedeelde studenten op, laat de gebruiker er een kiezen en
    voegt die toe aan coach-, tribe- of gildegroep in .env.
    Geeft (succes: bool, token: str) terug.
    """
    all_students, token = _fetch_shared(token)
    if not all_students:
        print("Geen studenten gevonden.")
        return False, token

    ordered = sorted(all_students.keys())
    student_name, token = _pick_student_from_list(ordered, "Studenten met gedeeld portfolio", token)
    if not student_name:
        return False, token

    return _pick_env_group_and_add(student_name), token


def _get_received_invitations(token: str) -> "list[str]":
    """
    Haalt studenten op die een evaluatieverzoek naar mij hebben gestuurd
    via GET /portfolios/{OWN_PORTFOLIO_ID}/progress-review/invitations/received.
    Geeft een gesorteerde lijst van unieke studentnamen terug.
    """
    if not OWN_PORTFOLIO_ID:
        print("PORTFLOW_OWN_PORTFOLIO_ID is niet ingesteld in .env.")
        return []

    headers = {"accept": "*/*", "authorization": f"Bearer {token}", "user-agent": "Mozilla/5.0"}
    names: set = set()
    page = 1

    while True:
        response = request_with_retries(
            f"{BASE_URL}/portfolios/{OWN_PORTFOLIO_ID}/progress-review/invitations/received",
            headers,
            params={"page": page, "per_page": PER_PAGE},
        )
        if response in (None, "TOKEN_EXPIRED"):
            break
        if isinstance(response, requests.Response) and response.status_code == 404:
            break
        try:
            data = response.json()
        except Exception:
            break
        if not isinstance(data, list) or not data:
            break
        for item in data:
            requester = item.get("requester") or {}
            name = (requester.get("name") or requester.get("full_name") or "").strip()
            if name:
                names.add(name)
        if len(data) < PER_PAGE:
            break
        page += 1

    return sorted(names)


def _add_invited_student_to_env(token: str) -> tuple:
    """
    Toont studenten die een evaluatieverzoek instuurden, laat er een kiezen
    en voegt die toe aan een groep in .env.
    Geeft (succes: bool, token: str) terug.
    """
    names = _get_received_invitations(token)
    if not names:
        print("Geen studenten gevonden die een evaluatieverzoek hebben gestuurd.")
        return False, token

    student_name, token = _pick_student_from_list(names, "Studenten met openstaand evaluatieverzoek", token)
    if not student_name:
        return False, token

    return _pick_env_group_and_add(student_name), token


def _add_non_coach_invited_student_to_env(token: str) -> tuple:
    """
    Toont studenten die een evaluatieverzoek instuurden maar niet in
    PORTFLOW_COACH_STUDENTS_JSON staan, laat er een kiezen en voegt
    die toe aan een groep in .env.
    Geeft (succes: bool, token: str) terug.
    """
    all_names = _get_received_invitations(token)
    if not all_names:
        print("Geen studenten gevonden die een evaluatieverzoek hebben gestuurd.")
        return False, token

    known_lower = {
        n.lower() for n in (COACH_STUDENT_NAMES | GILDE_STUDENT_NAMES)
        if n != SEPARATOR_SENTINEL
    }
    names = [n for n in all_names if n.lower() not in known_lower]

    if not names:
        print("Alle studenten die evaluatieverzoeken stuurden staan al in PORTFLOW_COACH_STUDENTS_JSON of PORTFLOW_GILDE_STUDENTS_JSON.")
        return False, token

    student_name, token = _pick_student_from_list(
        names, "Studenten met evaluatieverzoek (niet jouw coach-student)", token
    )
    if not student_name:
        return False, token

    return _pick_env_group_and_add(student_name), token


def _enrich_env_from_excel(xlsx_path: Path) -> bool:
    """
    Vult semester, tribe en gilde aan voor studenten die al in een van de drie
    groepen staan (PORTFLOW_COACH_STUDENTS_JSON, PORTFLOW_TRIBE_STUDENTS_JSON,
    PORTFLOW_GILDE_STUDENTS_JSON), op basis van naam-matching met het Excel-bestand.
    Kiest geen coach of gilde — verwerkt alle rijen.
    """
    try:
        import openpyxl as _xl
    except ImportError:
        print("Fout: openpyxl is niet geïnstalleerd. Voer 'pip install openpyxl' uit.")
        return False

    wb = _xl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(raw_headers) if h}

    # Bouw lookup: naam (lowercase) → velden uit Excel
    excel_by_name: dict = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[col["Nummer"]] is None:
            continue
        name = str(r[col["Naam"]] or "").strip()
        if not name:
            continue
        excel_by_name[name.lower()] = {
            "semester": _extract_semester_number(str(r[col["Semester"]] or "")),
            "tribe":    _extract_tribe_from_project(str(r[col["Project"]] or "")),
            "gilde":    str(r[col["Gilde voorkeur"]] or "").strip(),
            "coach":    str(r[col["Coach"]] or "").strip(),
        }

    if not excel_by_name:
        print("Geen studentrijen gevonden in het Excel-bestand.")
        return False

    _env_keys = [
        ("PORTFLOW_COACH_STUDENTS_JSON", _reload_coach_students),
        ("PORTFLOW_TRIBE_STUDENTS_JSON", _reload_tribe_students),
        ("PORTFLOW_GILDE_STUDENTS_JSON", _reload_gilde_students),
    ]

    total_updated = 0
    for env_key, reload_fn in _env_keys:
        existing_raw = os.getenv(env_key, "").strip()
        try:
            existing_list: list = json.loads(re.sub(r",\s*([\]\}])", r"\1", existing_raw)) if existing_raw else []
            if not isinstance(existing_list, list):
                existing_list = []
        except Exception:
            existing_list = []

        updated = 0
        result = []
        for entry in existing_list:
            if not isinstance(entry, dict) or entry.get("separator"):
                result.append(entry)
                continue
            key = (entry.get("name") or "").lower()
            if key in excel_by_name:
                xl = excel_by_name[key]
                entry = dict(entry)
                entry["semester"] = xl["semester"]
                entry["tribe"]    = xl["tribe"]
                entry["gilde"]    = xl["gilde"]
                entry["coach"]    = xl["coach"]
                updated += 1
            result.append(entry)

        if updated:
            result = _group_by_tribe([e for e in result if not e.get("separator")])
            parts = ["["]
            for i, e in enumerate(result):
                comma = "," if i < len(result) - 1 else ""
                parts.append(f"    {json.dumps(e, ensure_ascii=False)}{comma}")
            parts.append("    ]")
            json_str = "\n".join(parts)
            _write_env_key(env_key, json_str)
            os.environ[env_key] = json_str
            reload_fn()
            total_updated += updated

        print(f"  {env_key}: {updated} student(en) bijgewerkt.")

    print(f"  Totaal: {total_updated} student(en) aangevuld vanuit Excel.")
    return total_updated > 0


def _import_gilde_from_excel(xlsx_path: Path) -> bool:
    """
    Interactief: kies een gilde uit het Excel-bestand en voeg de studenten toe
    aan PORTFLOW_GILDE_STUDENTS_JSON. Studenten kunnen meerdere gildes hebben
    (gescheiden door '/').
    """
    try:
        import openpyxl as _xl
    except ImportError:
        print("Fout: openpyxl is niet geïnstalleerd. Voer 'pip install openpyxl' uit.")
        return False

    wb = _xl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(raw_headers) if h}

    rows = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if row[col["Nummer"]] is not None
    ]

    # Verzamel alle unieke gildes (split op '/')
    alle_gildes: set = set()
    for r in rows:
        raw_gilde = str(r[col["Gilde voorkeur"]] or "").strip()
        for g in raw_gilde.split("/"):
            g = g.strip()
            if g:
                alle_gildes.add(g)

    if not alle_gildes:
        print("Geen gildes gevonden in het Excel-bestand.")
        return False

    _HL = "\033[48;2;100;160;220m\033[1m"
    _RS = "\033[0m"

    gildes = sorted(alle_gildes)
    gilde_counts = {
        gilde: sum(
            1 for r in rows
            if any(g.strip() == gilde for g in str(r[col["Gilde voorkeur"]] or "").split("/"))
        )
        for gilde in gildes
    }

    def _print_gilde_menu(highlight_idx: str | None = None) -> None:
        print("\nBeschikbare gildes:")
        for i, gilde in enumerate(gildes, 1):
            label = f"{i}) {gilde}  ({gilde_counts[gilde]} studenten)"
            if str(i) == highlight_idx:
                print(f"  {_HL}{label}{_RS}")
            else:
                print(f"  {label}")
        print()

    _print_gilde_menu()
    raw = input("Kies een gilde (volgnummer of naam): ").strip()

    # Bepaal gekozen gilde
    if raw.isdigit() and 1 <= int(raw) <= len(gildes):
        chosen_gilde = gildes[int(raw) - 1]
        chosen_idx = raw
    else:
        matches = [g for g in gildes if raw.lower() == g.lower()]
        if not matches:
            matches = [g for g in gildes if raw.lower() in g.lower()]
        if len(matches) == 1:
            chosen_gilde = matches[0]
            chosen_idx = str(gildes.index(chosen_gilde) + 1)
        elif len(matches) > 1:
            print(f"Meerdere gildes gevonden: {', '.join(matches)}. Wees specifieker.")
            return False
        else:
            print("Gilde niet gevonden.")
            return False

    # Herdruk menu met highlight op gekozen gilde
    print(f"\033[{len(gildes) + 3}A\033[J", end="")
    _print_gilde_menu(chosen_idx)
    print(f"Kies een gilde (volgnummer of naam): {raw}")

    gilde_rows = [
        r for r in rows
        if any(
            g.strip() == chosen_gilde
            for g in str(r[col["Gilde voorkeur"]] or "").split("/")
        )
    ]
    print(f"\n{len(gilde_rows)} studenten gevonden voor gilde '{chosen_gilde}'.")

    new_entries = _group_by_tribe([
        {
            "name":       str(r[col["Naam"]] or "").strip(),
            "start_date": None,
            "semester":   _extract_semester_number(str(r[col["Semester"]] or "")),
            "tribe":      _extract_tribe_from_project(str(r[col["Project"]] or "")),
            "gilde":      str(r[col["Gilde voorkeur"]] or "").strip(),
            "coach":      str(r[col["Coach"]] or "").strip(),
        }
        for r in gilde_rows
        if str(r[col["Naam"]] or "").strip()
    ])

    existing_raw = os.getenv("PORTFLOW_GILDE_STUDENTS_JSON", "").strip()
    try:
        existing_list: list = json.loads(re.sub(r",\s*([\]\}])", r"\1", existing_raw)) if existing_raw else []
        if not isinstance(existing_list, list):
            existing_list = []
    except Exception:
        existing_list = []

    existing_names = {
        (e.get("name") or "").lower()
        for e in existing_list
        if isinstance(e, dict) and not e.get("separator")
    }
    to_add = [e for e in new_entries if not e.get("separator") and (e.get("name") or "").lower() not in existing_names]
    already = len(new_entries) - len([e for e in new_entries if e.get("separator")]) - len(to_add)

    if already:
        print(f"  {already} student(en) stonden al in PORTFLOW_GILDE_STUDENTS_JSON, worden overgeslagen.")
    if not to_add:
        print("  Geen nieuwe studenten toe te voegen.")
        return False

    # Voeg toe en sorteer volledige lijst op tribe + naam
    all_entries = [e for e in existing_list if not e.get("separator")] + to_add
    combined = _group_by_tribe(all_entries)

    parts = ["["]
    for i, e in enumerate(combined):
        comma = "," if i < len(combined) - 1 else ""
        parts.append(f"    {json.dumps(e, ensure_ascii=False)}{comma}")
    parts.append("    ]")
    json_str = "\n".join(parts)

    _write_env_key("PORTFLOW_GILDE_STUDENTS_JSON", json_str)
    os.environ["PORTFLOW_GILDE_STUDENTS_JSON"] = json_str
    _reload_gilde_students()
    print(f"  {len(to_add)} student(en) toegevoegd aan PORTFLOW_GILDE_STUDENTS_JSON.")
    return True


def _import_students_from_excel(xlsx_path: Path) -> bool:
    """
    Interactief: kies een coach, kies een importmodus en schrijf het resultaat
    naar .env als PORTFLOW_COACH_STUDENTS_JSON.
    Geeft True terug bij succes, False bij annulering of fout.
    """
    try:
        import openpyxl as _xl
    except ImportError:
        print("Fout: openpyxl is niet geïnstalleerd. Voer 'pip install openpyxl' uit.")
        return False

    wb = _xl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(raw_headers) if h}

    rows = [
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if row[col["Nummer"]] is not None
    ]

    # Unieke coaches, gesorteerd
    coaches = sorted({str(r[col["Coach"]] or "").strip() for r in rows} - {""})
    if not coaches:
        print("Geen coaches gevonden in het Excel-bestand.")
        return False

    print("\nBeschikbare coaches:")
    for i, coach in enumerate(coaches, 1):
        n = sum(1 for r in rows if str(r[col["Coach"]] or "").strip() == coach)
        print(f"  {i}) {coach}  ({n} studenten)")
    print()

    raw = input("Kies een coach (volgnummer of naam): ").strip()
    if raw.isdigit() and 1 <= int(raw) <= len(coaches):
        chosen = coaches[int(raw) - 1]
    else:
        matches = [c for c in coaches if raw.lower() in c.lower()]
        if len(matches) == 1:
            chosen = matches[0]
        elif len(matches) > 1:
            print(f"Meerdere coaches gevonden: {', '.join(matches)}. Wees specifieker.")
            return False
        else:
            print("Coach niet gevonden.")
            return False

    coach_rows = [r for r in rows if str(r[col["Coach"]] or "").strip() == chosen]
    print(f"\n{len(coach_rows)} studenten gevonden voor coach '{chosen}'.")

    # Bouw Excel-lookup: genormaliseerde naam → velden uit Excel
    def _excel_entry(r: tuple) -> dict:
        return {
            "name":       str(r[col["Naam"]] or "").strip(),
            "start_date": None,
            "semester":   _extract_semester_number(str(r[col["Semester"]] or "")),
            "tribe":      _extract_tribe_from_project(str(r[col["Project"]] or "")),
            "gilde":      str(r[col["Gilde voorkeur"]] or "").strip(),
            "coach":      chosen,
        }

    excel_by_name = {
        str(r[col["Naam"]] or "").strip().lower(): _excel_entry(r)
        for r in coach_rows
        if str(r[col["Naam"]] or "").strip()
    }

    # Huidige .env-waarde als lijst van dicts (inclusief separators)
    existing_raw = os.getenv("PORTFLOW_COACH_STUDENTS_JSON", "").strip()
    try:
        existing_list: list = json.loads(re.sub(r",\s*([\]\}])", r"\1", existing_raw)) if existing_raw else []
        if not isinstance(existing_list, list):
            existing_list = []
    except Exception:
        existing_list = []

    # Importmodus kiezen
    print("\nWat wil je doen?")
    print("  1) Veilig: vul alleen informatie aan voor studenten die al in .env staan")
    print("  2) Veilig: vul PORTFLOW_COACH_STUDENTS_JSON aan met studenten die wel in Excel,")
    print("             maar nog niet in .env staan")
    print("  3) Gevaarlijk: vervang PORTFLOW_COACH_STUDENTS_JSON door de inhoud van het excelbestand")
    print()
    mode = input("Keuze (1/2/3 of 'q' om te annuleren): ").strip()
    if mode.lower() == "q" or mode not in ("1", "2", "3"):
        print("Geannuleerd.")
        return False

    if mode == "1":
        # Update bestaande studenten: overschrijf semester/tribe/gilde uit Excel; behoud start_date
        updated = 0
        result = []
        for entry in existing_list:
            if not isinstance(entry, dict) or entry.get("separator"):
                result.append(entry)
                continue
            key = (entry.get("name") or "").lower()
            if key in excel_by_name:
                xl = excel_by_name[key]
                entry = dict(entry)
                entry["semester"] = xl["semester"]
                entry["tribe"]    = xl["tribe"]
                entry["gilde"]    = xl["gilde"]
                entry["coach"]    = xl["coach"]
                updated += 1
            result.append(entry)
        entries = _group_by_tribe([e for e in result if not e.get("separator")])
        print(f"  {updated} bestaande student(en) bijgewerkt.")

    elif mode == "2":
        # Voeg studenten toe die nog niet in .env staan
        existing_names = {
            (e.get("name") or "").lower()
            for e in existing_list
            if isinstance(e, dict) and not e.get("separator")
        }
        raw_new = [xl for key, xl in excel_by_name.items() if key not in existing_names]
        all_entries = [e for e in existing_list if not e.get("separator")] + raw_new
        entries = _group_by_tribe(all_entries)
        print(f"  {len(raw_new)} nieuwe student(en) toegevoegd (gegroepeerd op tribe).")

    else:  # mode == "3"
        ans = input(
            "Dit vervangt de volledige lijst in .env. Weet je het zeker? [j/N] "
        ).strip().lower()
        if ans not in ("j", "y", "ja", "yes"):
            print("Geannuleerd.")
            return False
        entries = _group_by_tribe([xl for xl in excel_by_name.values() if xl["name"]])
        n_students = sum(1 for e in entries if not e.get("separator"))
        print(f"  {n_students} studenten worden weggeschreven, gegroepeerd op tribe (volledige vervanging).")

    # Schrijf als meerregelige JSON (zelfde stijl als huidige .env)
    parts = ["["]
    for i, e in enumerate(entries):
        comma = "," if i < len(entries) - 1 else ""
        parts.append(f"    {json.dumps(e, ensure_ascii=False)}{comma}")
    parts.append("    ]")
    json_str = "\n".join(parts)

    _write_env_key("PORTFLOW_COACH_STUDENTS_JSON", json_str)
    os.environ["PORTFLOW_COACH_STUDENTS_JSON"] = json_str
    _reload_coach_students()
    print(f"  Klaar. PORTFLOW_COACH_STUDENTS_JSON bijgewerkt in .env.")
    return True

def request_with_retries(url, headers, params=None, max_attempts=3):
    attempt = 0
    while attempt < max_attempts:
        try:
            response = requests.get(url, headers=headers, params=params, timeout=15)
            response_body = response.text

            if response.status_code == 401:
                log_api_debug(url, params, status_code=401, body=response_body)
                return "TOKEN_EXPIRED"

            if response.status_code == 404:
                # Missing resources (e.g. student without linked portfolio) should not be retried.
                log_api_debug(url, params, status_code=404, body=response_body)
                return response

            response.raise_for_status()
            log_api_debug(url, params, status_code=response.status_code, body=response_body)
            return response

        except requests.exceptions.RequestException as e:
            attempt += 1
            log_api_debug(url, params, error=str(e))
            print(f"Request failed ({attempt}/{max_attempts}): {e}")
            if attempt < max_attempts:
                print("Retrying in 5 seconds...")
                time.sleep(5)
            else:
                print("3 failed attempts. Continuing...")
                return None

# ------------------------
# Student fetching (paginated)
# ------------------------

def get_shared_collections(token):
    headers = {
        "accept": "*/*",
        "authorization": f"Bearer {token}",
        "user-agent": "Mozilla/5.0"
    }

    all_items = []
    page = 1

    while True:
        response = request_with_retries(
            f"{BASE_URL}/shares/shared-with-me",
            headers,
            params={
                "order_by": "created_at",
                "order_direction": "desc",
                "page": page,
                "per_page": PER_PAGE
            }
        )

        if response in (None, "TOKEN_EXPIRED"):
            return response

        data = response.json()
        observe_schema(data, "shares.shared_with_me.response")
        if not data:
            break

        all_items.extend(data)

        if len(data) < 20:
            break

        page += 1

    return all_items

def get_students_from_section(token, section_id):
    headers = {
        "accept": "*/*",
        "authorization": f"Bearer {token}",
        "user-agent": "Mozilla/5.0"
    }

    students = {}
    page = 1

    while True:
        response = request_with_retries(
            f"{BASE_URL}/dashboard",
            headers,
            params={
                "section_id": section_id,
                "page": page,
                "per_page": PER_PAGE
            }
        )

        if response in (None, "TOKEN_EXPIRED"):
            return response

        data = response.json()
        observe_schema(data, "dashboard.response")
        page_students = data.get("students", [])

        if not page_students:
            break

        for student in page_students:
            name = student["name"]
            portfolio_id = student["portfolio_id"]
            if name not in students:
                students[name] = {
                    "student_id": student["id"],
                    "portfolio_ids": set()
                }
            students[name]["portfolio_ids"].add(portfolio_id)

        if len(page_students) < PER_PAGE:
            break

        page += 1

    return students

def extract_students(shared_items):
    students = {}
    for item in shared_items:
        inviter = item.get("inviter")
        if not inviter or inviter.get("current_role") != "student":
            continue

        name = inviter["name"]
        portfolio_id = item["portfolio_id"]

        if name not in students:
            students[name] = {
                "student_id": inviter["id"],
                "portfolio_ids": set()
            }

        students[name]["portfolio_ids"].add(portfolio_id)

    return students

# ------------------------
# Portfolio & feedback fetching
# ------------------------

def get_sent_invitations(token, portfolio_id) -> dict:
    """
    Haalt openstaande review requests op die vanuit dit portfolio zijn verzonden.
    Geeft een dict terug: review_request_id → reviewer_name.

    Gebruikt: GET /portfolios/{id}/progress-review/invitations/sent
    """
    headers = {"accept": "*/*", "authorization": f"Bearer {token}", "user-agent": "Mozilla/5.0"}
    response = request_with_retries(
        f"{BASE_URL}/portfolios/{portfolio_id}/progress-review/invitations/sent",
        headers,
        params={"per_page": PER_PAGE},
    )
    if response in (None, "TOKEN_EXPIRED"):
        return {}
    if isinstance(response, requests.Response) and response.status_code == 404:
        return {}
    try:
        data = response.json()
    except Exception:
        return {}
    if not isinstance(data, list):
        return {}

    mapping: dict = {}
    for item in data:
        rid = item.get("review_request_id")
        if not rid:
            continue
        reviewer_obj = item.get("reviewer") or {}
        name = (
            reviewer_obj.get("name")
            or reviewer_obj.get("full_name")
            or item.get("reviewer_name")
        )
        if name:
            mapping[rid] = name
    return mapping


def get_goals(token, portfolio_id):
    headers = {"accept": "*/*", "authorization": f"Bearer {token}"}
    response = request_with_retries(
        f"{BASE_URL}/portfolios/{portfolio_id}/goals",
        headers,
        params={"page": 1, "per_page": PER_PAGE}
    )
    if response in (None, "TOKEN_EXPIRED"):
        return response
    if isinstance(response, requests.Response) and response.status_code == 404:
        return "NOT_FOUND"
    data = response.json()
    observe_schema(data, "portfolios.goals.response")
    return data

def get_feedback(token, portfolio_id, goal_id):
    headers = {"accept": "*/*", "authorization": f"Bearer {token}"}
    feedback_items = []
    page = 1

    while True:
        response = request_with_retries(
            f"{BASE_URL}/portfolios/{portfolio_id}/goals/{goal_id}/feedback-items",
            headers,
            params={"page": page, "per_page": PER_PAGE}
        )

        if isinstance(response, requests.Response) and response.status_code == 404:
            return "NOT_FOUND"

        if response is None:
            break

        if response == "TOKEN_EXPIRED":
            return "TOKEN_EXPIRED"

        data = response.json()
        observe_schema(data, "portfolios.goals.feedback_items.response")
        if not data:
            break

        feedback_items.extend(data)

        if len(data) < PER_PAGE:
            break

        page += 1

    return feedback_items

def resolve_level(evaluation):
    level_id = evaluation.get("level")
    if not level_id:
        return None

    for lvl in evaluation.get("level_set", []):
        if lvl["id"] == level_id:
            label = lvl["label"]
            return "0" if label == "Startniveau" else label

    return None


def pending_reason(item, evaluation):
    if not evaluation:
        return "missing_evaluation"

    review_request_scored = evaluation.get("review_request_scored")
    level = evaluation.get("level")
    role = item.get("role")

    if review_request_scored is False:
        return "review_request_scored_false"
    if level in (None, ""):
        return "missing_level"
    if role == "self" and review_request_scored is not True:
        return "self_not_confirmed_scored"
    return None

def resolve_reviewer_name(item, evaluation=None):
    # Best-effort extraction of the assessor/reviewer name from the API payload.
    candidates = [
        (item.get("user") or {}).get("name"),
        (item.get("user") or {}).get("full_name"),
        (item.get("author") or {}).get("name"),
        (item.get("creator") or {}).get("name"),
        (item.get("reviewer") or {}).get("name"),
        (item.get("evaluator") or {}).get("name"),
        (item.get("created_by") or {}).get("name"),
        (item.get("createdBy") or {}).get("name"),
        item.get("user_name"),
        item.get("author_name"),
        item.get("creator_name"),
        item.get("reviewer_name"),
        item.get("evaluator_name"),
        item.get("created_by_name"),
        item.get("createdByName"),
    ]
    for value in candidates:
        if isinstance(value, str) and value.strip():
            return value.strip()

    hint_tokens = (
        "review",
        "evaluator",
        "assessor",
        "author",
        "coach",
        "teacher",
        "created_by",
        "createdby",
        "creator",
        "user",
        "inviter",
        "sender",
        "owner",
        "by",
    )
    name_keys = ("name", "full_name", "fullname", "display_name", "displayname")

    def walk(value, in_context: bool = False):
        if isinstance(value, dict):
            for k, v in value.items():
                k_lower = str(k).lower()
                next_context = in_context or any(tok in k_lower for tok in hint_tokens)

                if k_lower in name_keys and next_context and isinstance(v, str) and v.strip():
                    return v.strip()

                if "name" in k_lower and next_context and isinstance(v, str) and v.strip():
                    return v.strip()

                found = walk(v, next_context)
                if found:
                    return found
        elif isinstance(value, list):
            for entry in value:
                found = walk(entry, in_context)
                if found:
                    return found
        return None

    found = walk(item)
    if found:
        return found
    if evaluation is not None:
        return walk(evaluation)
    return None


def parse_datetime_value(raw_value):
    if raw_value in (None, ""):
        return None

    if isinstance(raw_value, datetime):
        return raw_value

    # Handle numeric epoch values (seconds or milliseconds).
    if isinstance(raw_value, (int, float)):
        timestamp = float(raw_value)
        if timestamp > 1e12:
            timestamp /= 1000.0
        try:
            return datetime.fromtimestamp(timestamp)
        except (OverflowError, OSError, ValueError):
            return None

    if not isinstance(raw_value, str):
        return None

    text = raw_value.strip()
    if not text:
        return None

    parse_candidates = [text]
    if text.endswith("Z"):
        parse_candidates.insert(0, text[:-1] + "+00:00")

    for candidate in parse_candidates:
        try:
            return datetime.fromisoformat(candidate)
        except ValueError:
            pass

    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue

    return None


def format_short_date(raw_value):
    dt = parse_datetime_value(raw_value)
    if dt is None:
        return None
    return f"{dt.day}-{dt.strftime('%b').lower()}-{dt.strftime('%y')}"


def date_in_selected_semester(raw_value, semester_scope: str, override_start=None) -> bool:
    if semester_scope == "all":
        return True

    if isinstance(raw_value, datetime):
        dt = raw_value
    else:
        dt = parse_datetime_value(raw_value)
    if dt is None:
        return False

    if semester_scope == "sep25_jan26":
        return SEMESTER_SEP25_JAN26_START <= dt.date() <= SEMESTER_SEP25_JAN26_END

    start = override_start if override_start is not None else CURRENT_SEMESTER_START
    return dt.date() >= start


def resolve_evaluation_date(item, evaluation=None):
    # Prefer explicit assessment timestamps before generic create/update fields.
    date_keys = (
        "submitted_at",
        "submittedat",
        "date",
        "evaluated_at",
        "evaluatedat",
        "evaluation_date",
        "evaluationdate",
        "graded_at",
        "gradedat",
        "reviewed_at",
        "reviewedat",
        "assessed_at",
        "assessedat",
        "created_at",
        "createdat",
        "updated_at",
        "updatedat",
    )

    for source in (evaluation, item):
        if not isinstance(source, dict):
            continue
        for raw_key, raw_value in source.items():
            normalized_key = str(raw_key).lower().replace("_", "")
            if normalized_key in date_keys:
                parsed = parse_datetime_value(raw_value)
                if parsed:
                    return parsed

    hint_tokens = ("evaluat", "grade", "review", "assess", "created", "updated", "submit")
    date_tokens = ("date", "time", "_at", "at")

    def walk(value, in_context=False):
        if isinstance(value, dict):
            for k, v in value.items():
                key = str(k).lower()
                next_context = in_context or any(tok in key for tok in hint_tokens)
                looks_like_date_key = any(tok in key for tok in date_tokens)

                if looks_like_date_key and (next_context or key in date_keys):
                    parsed = parse_datetime_value(v)
                    if parsed:
                        return parsed

                found = walk(v, next_context)
                if found:
                    return found

        elif isinstance(value, list):
            for entry in value:
                found = walk(entry, in_context)
                if found:
                    return found

        return None

    return walk(item) or walk(evaluation)

def collect_results(
    token,
    student_name,
    student_data,
    semester_scope="current",
    override_start=None,
    include_ungraded=False,
    inaccessible_names=None,
):
    results = []
    any_accessible = False
    portfolio_ids = list(student_data["portfolio_ids"])

    for portfolio_id in portfolio_ids:
        goals = get_goals(token, portfolio_id)
        if goals == "TOKEN_EXPIRED":
            return "TOKEN_EXPIRED"
        if goals == "NOT_FOUND":
            print(
                f"[404] No linked portfolio for {student_name} "
                f"(student_id={student_data.get('student_id')}, portfolio_id={portfolio_id}). Skipping."
            )
            continue
        if goals is None:
            print(
                f"[ERROR] Failed to fetch goals for {student_name} "
                f"(student_id={student_data.get('student_id')}, portfolio_id={portfolio_id}). Skipping."
            )
            continue
        any_accessible = True

        # Haal reviewer-namen op voor openstaande review requests (self-evaluaties).
        # Alleen nodig als we pending items tonen.
        rid_to_reviewer: dict = get_sent_invitations(token, portfolio_id) if include_ungraded else {}

        for goal in goals:
            goal_id = goal["id"]
            goal_name = goal["name"]

            feedback_items = get_feedback(token, portfolio_id, goal_id)
            if feedback_items == "TOKEN_EXPIRED":
                return "TOKEN_EXPIRED"
            if feedback_items == "NOT_FOUND":
                print(
                    f"[404] Feedback endpoint not found for {student_name} "
                    f"(portfolio_id={portfolio_id}, goal_id={goal_id}). Skipping goal."
                )
                continue

            for item in feedback_items:
                observe_schema(item, "portfolios.goals.feedback_items.item")
                if item.get("type") != "criterion_evaluation":
                    continue

                evaluation = item.get("evaluation")
                pending = pending_reason(item, evaluation)

                if not evaluation and not include_ungraded:
                    log_pending_debug_event({
                        "student_name": student_name,
                        "portfolio_id": portfolio_id,
                        "goal_id": goal_id,
                        "goal_name": goal_name,
                        "role": item.get("role"),
                        "decision": "skip",
                        "reason": "missing_evaluation",
                    })
                    continue

                if evaluation:
                    observe_schema(evaluation, "portfolios.goals.feedback_items.item.evaluation")

                evaluation_datetime = resolve_evaluation_date(item, evaluation)
                # Only filter by date if a date was actually found; items without
                # a resolvable date are included (we cannot prove they are outside
                # the semester and dropping them would hide real evaluations).
                if evaluation_datetime is not None and not date_in_selected_semester(evaluation_datetime, semester_scope, override_start):
                    log_pending_debug_event({
                        "student_name": student_name,
                        "portfolio_id": portfolio_id,
                        "goal_id": goal_id,
                        "goal_name": goal_name,
                        "role": item.get("role"),
                        "decision": "skip",
                        "reason": "outside_semester",
                        "detected_date": evaluation_datetime.isoformat() if evaluation_datetime else None,
                        "item_created_at": item.get("created_at"),
                        "item_updated_at": item.get("updated_at"),
                        "eval_created_at": (evaluation or {}).get("created_at"),
                        "eval_submitted_at": (evaluation or {}).get("submitted_at"),
                        "pending_reason": pending,
                        "review_request_scored": (evaluation or {}).get("review_request_scored"),
                        "level": (evaluation or {}).get("level"),
                    })
                    continue

                # Keep historical behavior for non-table modes: self evaluations are hidden.
                if item.get("role") == "self" and not include_ungraded:
                    log_pending_debug_event({
                        "student_name": student_name,
                        "portfolio_id": portfolio_id,
                        "goal_id": goal_id,
                        "goal_name": goal_name,
                        "role": item.get("role"),
                        "decision": "skip",
                        "reason": "self_role_hidden",
                        "pending_reason": pending,
                    })
                    continue

                level = resolve_level(evaluation)
                is_pending = pending is not None

                if is_pending:
                    if not include_ungraded:
                        log_pending_debug_event({
                            "student_name": student_name,
                            "portfolio_id": portfolio_id,
                            "goal_id": goal_id,
                            "goal_name": goal_name,
                            "role": item.get("role"),
                            "decision": "skip",
                            "reason": "pending_hidden",
                            "pending_reason": pending,
                            "review_request_scored": (evaluation or {}).get("review_request_scored"),
                            "level": (evaluation or {}).get("level"),
                            "review_request_title": (evaluation or {}).get("review_request_title"),
                        })
                        continue

                    # In table mode, pending items are shown as '?'.
                    evaluation_text = "?"
                else:
                    # Avoid exposing scored self-evaluations in table mode.
                    if item.get("role") == "self":
                        log_pending_debug_event({
                            "student_name": student_name,
                            "portfolio_id": portfolio_id,
                            "goal_id": goal_id,
                            "goal_name": goal_name,
                            "role": item.get("role"),
                            "decision": "skip",
                            "reason": "self_scored_hidden",
                            "review_request_scored": (evaluation or {}).get("review_request_scored"),
                            "level": (evaluation or {}).get("level"),
                        })
                        continue

                    if level is None:
                        log_pending_debug_event({
                            "student_name": student_name,
                            "portfolio_id": portfolio_id,
                            "goal_id": goal_id,
                            "goal_name": goal_name,
                            "role": item.get("role"),
                            "decision": "skip",
                            "reason": "missing_level_unexpected",
                            "review_request_scored": (evaluation or {}).get("review_request_scored"),
                        })
                        continue

                    evaluation_text = str(level)

                reviewer_name = resolve_reviewer_name(item, evaluation)
                initials = name_to_initials(reviewer_name) if reviewer_name else ""

                details = []
                if (not is_pending) and initials:
                    details.append(initials)
                evaluation_date = format_short_date(evaluation_datetime)
                if (not is_pending) and evaluation_date:
                    details.append(evaluation_date)

                if details:
                    evaluation_text = f"{level} ({', '.join(details)})"

                _pending_role = item.get("role")
                if _pending_role == "self" and is_pending:
                    # evaluation.reviewer is the student themselves on self-eval items.
                    # Look up the actual assigned assessor via the sent-invitations map.
                    _rid = (evaluation or {}).get("review_request_id")
                    _ev_reviewer = rid_to_reviewer.get(_rid) if _rid else None
                else:
                    _ev_reviewer = (
                        None if _pending_role == "self"
                        else ((evaluation or {}).get("reviewer") or {}).get("name")
                    )
                results.append({
                    "student_name": student_name,
                    "goal_name": goal_name,
                    "evaluation": evaluation_text,
                    "submitted_at_iso": evaluation_datetime.isoformat() if evaluation_datetime else None,
                    "pending_detail": {
                        "reviewer": _ev_reviewer,
                        "date": evaluation_date,
                        "title": (evaluation or {}).get("review_request_title"),
                    } if is_pending else None,
                })

                log_pending_debug_event({
                    "student_name": student_name,
                    "portfolio_id": portfolio_id,
                    "goal_id": goal_id,
                    "goal_name": goal_name,
                    "role": item.get("role"),
                    "decision": "include",
                    "rendered": evaluation_text,
                    "pending_reason": pending,
                    "review_request_scored": (evaluation or {}).get("review_request_scored"),
                    "level": (evaluation or {}).get("level"),
                    "review_request_title": (evaluation or {}).get("review_request_title"),
                    "submitted_at": (evaluation or {}).get("submitted_at"),
                })

            # Detect pending coach evaluations that are not yet visible in the
            # API because the assigned coach hasn't responded yet (e.g., review
            # requests sent to a different coach).  When a self-evaluation for a
            # review_request_id is within the current semester but no coach
            # feedback item exists for that same review_request_id, add a
            # synthetic '?' so the pending state is visible in the table.
            if include_ungraded:
                self_by_rid: dict = {}    # rid -> (item, evaluation)
                reviewer_rids: set = set()  # rids with any coach/assessor response

                for _item in feedback_items:
                    if _item.get("type") != "criterion_evaluation":
                        continue
                    _ev = _item.get("evaluation") or {}
                    _rid = _ev.get("review_request_id")
                    if not _rid:
                        continue
                    _role = _item.get("role")
                    if _role in ("coach", "assessor"):
                        reviewer_rids.add(_rid)
                    elif _role == "self":
                        _dt = resolve_evaluation_date(_item, _ev)
                        _in = _dt is None or date_in_selected_semester(_dt, semester_scope, override_start)
                        if _in:
                            self_by_rid.setdefault(_rid, (_item, _ev))

                for _rid, (_sitem, _sev) in self_by_rid.items():
                    if _rid in reviewer_rids:
                        continue
                    _self_dt = resolve_evaluation_date(_sitem, _sev)
                    results.append({
                        "student_name": student_name,
                        "goal_name": goal_name,
                        "evaluation": "?",
                        "submitted_at_iso": _self_dt.isoformat() if _self_dt else None,
                        "pending_detail": {
                            "reviewer": rid_to_reviewer.get(_rid),
                            "date": format_short_date(_self_dt),
                            "title": _sev.get("review_request_title"),
                        },
                    })
                    log_pending_debug_event({
                        "student_name": student_name,
                        "portfolio_id": portfolio_id,
                        "goal_id": goal_id,
                        "goal_name": goal_name,
                        "role": "coach",
                        "decision": "include",
                        "rendered": "?",
                        "reason": "coach_pending_inferred",
                        "review_request_id": _rid,
                        "review_request_title": _sev.get("review_request_title"),
                        "self_submitted_at": _sev.get("submitted_at"),
                    })

    if portfolio_ids and not any_accessible and inaccessible_names is not None:
        inaccessible_names.add(student_name)

    return results


def print_week_barchart(results, override_start=None):
    import math as _math
    from datetime import timedelta as _td
    sem_start = override_start if override_start is not None else CURRENT_SEMESTER_START
    sem_end   = CURRENT_SEMESTER_END

    # Weeks needed to cover until semester end
    num_weeks = _math.ceil(((sem_end - sem_start).days + 1) / 7)
    if num_weeks <= 0:
        return

    _counted_levels = {"1", "2", "3", "?"}
    counts = [0] * num_weeks
    for r in results:
        level = extract_level_short(r.get("evaluation", ""))
        if level not in _counted_levels:
            continue
        raw = r.get("submitted_at_iso")
        if not raw:
            continue
        dt = parse_datetime_value(raw)
        if dt is None:
            continue
        d = dt.date() if isinstance(dt, datetime) else dt
        if d < sem_start:
            continue
        week_idx = (d - sem_start).days // 7
        if 0 <= week_idx < num_weeks:
            counts[week_idx] += 1

    if not any(c > 0 for c in counts):
        return

    max_count = max(counts)
    COL_W = 5  # chars per week column: " ██  " or "     "

    # Y-axis label width: digits + 1 space before │
    y_w = len(str(max_count)) + 1

    # Week start dates (Monday of each week from sem_start)
    week_starts = [sem_start + _td(days=7 * i) for i in range(num_weeks)]

    # Alignment helpers
    #   bar rows:   "{y_w chars} │{bar data}"  → bars start at col y_w+2
    #   axis:       "{y_w+1 spaces}╠{'═'*...}╣"
    #   week/date:  "{y_w+2 spaces}{labels}"
    axis_prefix  = " " * (y_w + 1)
    label_prefix = " " * (y_w + 2)

    print()

    # Bar rows from max_count down to 1
    for row_val in range(max_count, 0, -1):
        y_label = f"{row_val:>{y_w}} │"
        row = "".join(" ██  " if c >= row_val else "     " for c in counts)
        print(y_label + row)

    # X-axis
    print(f"{axis_prefix}╠{'═' * (num_weeks * COL_W)}╣")

    # Week number labels: W1, W2, ... W21
    print(label_prefix + "".join(f"W{i + 1:<{COL_W - 1}}" for i in range(num_weeks)))

    # Date labels: day/month of the Monday of each week
    print(label_prefix + "".join(f"{ws.day}/{ws.month:<{COL_W - len(str(ws.day)) - 1}}" for ws in week_starts))


def print_student_evaluations(token, student_name, student_data, semester_scope="current", override_start=None):
    results = collect_results(token, student_name, student_data, semester_scope, override_start=override_start, include_ungraded=bool(ARGS.vraagtekens))
    if results == "TOKEN_EXPIRED":
        return "TOKEN_EXPIRED"

    # Splits in behaald en ingediend (pending)
    achieved: dict = {}   # goal -> list of formatted strings
    pending: dict = {}    # goal -> list of formatted strings

    for r in results:
        goal = r["goal_name"]
        ev = r["evaluation"]
        pd = r.get("pending_detail")

        if ev == "?" and pd:
            reviewer = pd.get("reviewer")
            date = pd.get("date")
            title = pd.get("title")
            initials = name_to_initials(reviewer) if reviewer else ""
            parts = []
            if initials:
                parts.append(initials)
            if date:
                parts.append(date)
            if title:
                short_title = title[:27] + "…" if len(title) > 28 else title
                parts.append(f'"{short_title}"')
            detail = ", ".join(parts) if parts else ""
            pending.setdefault(goal, []).append(f"? ({detail})" if detail else "?")
        else:
            achieved.setdefault(goal, []).append(ev)

    if achieved:
        print("Behaalde evaluaties:")
        for goal, parts in achieved.items():
            print(f"  {goal}: {', '.join(parts)}")
    else:
        print("Behaalde evaluaties: (geen)")

    if ARGS.vraagtekens:
        if pending:
            print("\nIngediende evaluaties (nog niet beoordeeld):")
            for goal, parts in pending.items():
                print(f"  {goal}: {', '.join(parts)}")
        else:
            print("\nIngediende evaluaties (nog niet beoordeeld): (geen)")

    print_week_barchart(results, override_start=override_start)
    return "OK"

# ------------------------
# CSV Export (wide format, ; separator)
# ------------------------

def extract_level_short(evaluation_text: str) -> str:
    """Strip reviewer/date details and return just the level value."""
    if " (" in evaluation_text:
        return evaluation_text.split(" (")[0]
    return evaluation_text


def print_coach_table(results, all_names=None, inaccessible_names=None, sem_map=None, tribe_map=None, gilde_map=None, coach_map=None):
    if not results and not all_names:
        print("No data to display.")
        return

    _sem_map   = sem_map   if sem_map   is not None else COACH_STUDENT_SEMESTER
    _tribe_map = tribe_map if tribe_map is not None else COACH_STUDENT_TRIBE
    _gilde_map = gilde_map if gilde_map is not None else COACH_STUDENT_GILDE
    _coach_map = coach_map  # None = geen Coach-kolom tonen

    # Achtergrondkleuren per groep (24-bit ANSI, passend bij de Open ICT kleurkaart)
    _BG_BLUE  = "\033[48;2;119;140;163m"   # blauw-grijs  (OC, KO, JKO, KPM)
    _BG_GREEN = "\033[48;2;141;164;122m"   # salie-groen  (PL, BD, SW)
    _BG_TERRA = "\033[48;2;190;145;110m"   # terra-cotta  (FO, PH, RE)
    _RESET    = "\033[0m"

    _GOAL_BG = {
        "Overzicht creëren":           _BG_BLUE,
        "Kritisch oordelen":           _BG_BLUE,
        "Juiste kennis ontwikkelen":   _BG_BLUE,
        "Kwalitatief Product Maken":   _BG_BLUE,
        "Plannen":                     _BG_GREEN,
        "Boodschap Delen":             _BG_GREEN,
        "Samenwerken":                 _BG_GREEN,
        "Flexibel opstellen":          _BG_TERRA,
        "Pro-actief handelen":         _BG_TERRA,
        "Reflecteren":                 _BG_TERRA,
    }

    # Build student -> goal -> list of short levels
    student_goals = {}
    for r in results:
        s = r["student_name"]
        g = r["goal_name"]
        level = extract_level_short(r["evaluation"])
        student_goals.setdefault(s, {}).setdefault(g, []).append(level)

    # Ensure all expected names are present (even with 0 evaluations); skip separators
    if all_names:
        for name in all_names:
            if name != SEPARATOR_SENTINEL:
                student_goals.setdefault(name, {})

    # Determine column widths
    col_widths = []
    for full_name, abbrev in GOAL_COLUMNS:
        width = len(abbrev)
        for goals in student_goals.values():
            cell = ", ".join(goals.get(full_name, []))
            width = max(width, len(cell))
        col_widths.append(width)

    name_width  = max(len("Naam"), max(len(s) for s in student_goals))
    sem_width   = max(len("Semester"), max((len(f"Semester {_sem_map.get(s, '')}".strip() if _sem_map.get(s) else "") for s in student_goals), default=0))
    tribe_width = max(len("Tribe"), max((len(_tribe_map.get(s, "")) for s in student_goals), default=0))
    gilde_width = max(len("Gilde"), max((len(_gilde_map.get(s, "")) for s in student_goals), default=0))
    coach_width = max(len("Coach"), max((len(_coach_map.get(s, "")) for s in student_goals), default=0)) if _coach_map is not None else 0

    # Bereken visuele breedte voor de scheidingslijn (zonder ANSI-codes)
    _plain_header = f"{'Naam':<{name_width}}"
    for (_, abbrev), w in zip(GOAL_COLUMNS, col_widths):
        _plain_header += f" | {abbrev:<{w}}"
    _plain_header += f" | {'Tribe':<{tribe_width}}"
    _plain_header += f" | {'Semester':<{sem_width}}"
    _plain_header += f" | {'Gilde':<{gilde_width}}"
    if _coach_map is not None:
        _plain_header += f" | {'Coach':<{coach_width}}"
    separator = "-" * len(_plain_header)

    # Gekleurde header
    header = f"{'Naam':<{name_width}}"
    for (full_name, abbrev), w in zip(GOAL_COLUMNS, col_widths):
        bg = _GOAL_BG[full_name]
        header += f" | {bg}{abbrev:<{w}}{_RESET}"
    header += f" | {'Tribe':<{tribe_width}}"
    header += f" | {'Semester':<{sem_width}}"
    header += f" | {'Gilde':<{gilde_width}}"
    if _coach_map is not None:
        header += f" | {'Coach':<{coach_width}}"
    print()
    print(header)
    print(separator)

    # Rows (in .env order if available, otherwise alphabetical)
    row_order = all_names if all_names else sorted(student_goals.keys())
    for student_name in row_order:
        if student_name == SEPARATOR_SENTINEL:
            print(separator)
            continue
        if student_name not in student_goals:
            continue
        goals = student_goals[student_name]
        display_name = anonymize_name(student_name) if ARGS.anoniem else student_name
        sem_raw = _sem_map.get(student_name, "")
        sem = f"Semester {sem_raw}" if sem_raw else ""
        tribe = _tribe_map.get(student_name, "")
        gilde = _gilde_map.get(student_name, "")
        coach = _coach_map.get(student_name, "") if _coach_map is not None else ""
        display_tribe = "******" if ARGS.anoniem else tribe
        display_sem = "******" if ARGS.anoniem else sem
        display_gilde = "******" if ARGS.anoniem else gilde
        display_coach = "******" if ARGS.anoniem else coach

        # Inaccessible student: portfolio niet zichtbaar voor deze coach
        if inaccessible_names and student_name in inaccessible_names:
            row = f"\033[2m{display_name:<{name_width}}\033[0m"
            for (full_name, _), w in zip(GOAL_COLUMNS, col_widths):
                bg = _GOAL_BG[full_name]
                row += f" | \033[2m{'n/b' if w >= 3 else '-':<{w}}\033[0m"
            row += f" | \033[2m{display_tribe:<{tribe_width}}\033[0m"
            row += f" | \033[2m{display_sem:<{sem_width}}\033[0m"
            row += f" | \033[2m{display_gilde:<{gilde_width}}\033[0m"
            if _coach_map is not None:
                row += f" | \033[2m{display_coach:<{coach_width}}\033[0m"
            print(row)
            continue

        _alert_goals = {"Overzicht creëren", "Kritisch oordelen", "Juiste kennis ontwikkelen", "Kwalitatief Product Maken"}
        _alert = all(not goals.get(fn) for fn, _ in GOAL_COLUMNS if fn in _alert_goals)
        _soft_goals = {"Plannen", "Boodschap Delen", "Samenwerken", "Flexibel opstellen", "Pro-actief handelen", "Reflecteren"}
        _soft_empty = sum(1 for fn, _ in GOAL_COLUMNS if fn in _soft_goals and not goals.get(fn))
        _soft_all   = _soft_empty == len(_soft_goals)   # alle 6 leeg → oranje
        _soft_most  = _soft_empty in (4, 5)             # 4 of 5 leeg → geel
        row = f"{display_name:<{name_width}}"
        for (full_name, _), w in zip(GOAL_COLUMNS, col_widths):
            cell = ", ".join(goals.get(full_name, []))
            if _alert and full_name in _alert_goals and not cell:
                row += f" | \033[41m{' ' * w}\033[0m"
            elif _soft_all and full_name in _soft_goals and not cell:
                row += f" | \033[43m{' ' * w}\033[0m"
            elif _soft_most and full_name in _soft_goals and not cell:
                row += f" | \033[33m{' ' * w}\033[0m"
            elif "?" in cell:
                row += f" | \033[104m{cell:<{w}}\033[0m"
            else:
                row += f" | {cell:<{w}}"
        row += f" | {display_tribe:<{tribe_width}}"
        row += f" | {display_sem:<{sem_width}}"
        row += f" | {display_gilde:<{gilde_width}}"
        if _coach_map is not None:
            row += f" | {display_coach:<{coach_width}}"
        print(row)

    if inaccessible_names and any(n in inaccessible_names for n in (all_names or []) if n != SEPARATOR_SENTINEL):
        print("  \033[2mn/b = portfolio niet zichtbaar (geen toegang)\033[0m")
    print()


def export_csv_wide(results):
    if not results:
        print("No data to export.")
        return

    all_goals = sorted(set(r["goal_name"] for r in results))
    students = {}

    for r in results:
        s = r["student_name"]
        g = r["goal_name"]

        students.setdefault(s, {goal: "" for goal in all_goals})

        if students[s][g]:
            students[s][g] += f", {r['evaluation']}"
        else:
            students[s][g] = r["evaluation"]

    with open("results.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Studentname"] + all_goals)

        for student, goal_data in students.items():
            writer.writerow([student] + [goal_data[g] for g in all_goals])

    print("CSV exported to results.csv")

# ------------------------
# Main loop (Ctrl+C safe)
# ------------------------


def _show_progress(current: int, total: int, label: str, bar_width: int = 25) -> None:
    filled = int(bar_width * current / total)
    arrow = ">" if filled < bar_width else "="
    bar = "=" * filled + arrow + " " * (bar_width - filled - 1)
    display_label = "" if current == total else label
    print(f"\r[{bar}] {current}/{total}  {display_label:<40}", end="", flush=True)


def _fetch_shared(token):
    """Haal alle gedeelde studenten op; herhaalt bij token-expiry."""
    while True:
        shared_items = get_shared_collections(token)
        if shared_items == "TOKEN_EXPIRED":
            print("Token verlopen, nieuw token ophalen...")
            token = get_bearer_token(force_prompt=True)
            continue
        return extract_students(shared_items), token


_migrate_env_add_coach_field()

# Bepaal het te gebruiken Excel-bestand (eenmalig bij opstarten)
_xlsx_candidates = _find_excel_with_student_columns() if ARGS.admin else []
if ARGS.admin and len(_xlsx_candidates) > 1:
    _HL = "\033[48;2;100;160;220m\033[1m"
    _RS = "\033[0m"
    print("\nMeerdere Excel-bestanden gevonden met de juiste kolomkoppen:")
    for _i, (_, _fn_c) in enumerate(_xlsx_candidates, 1):
        print(f"  {_i}) [{_fn_c}]")
    print()
    _raw_xls = input(f"Welk bestand wil je gebruiken? (1-{len(_xlsx_candidates)}): ").strip()
    if _raw_xls.isdigit() and 1 <= int(_raw_xls) <= len(_xlsx_candidates):
        _xlsx_idx = int(_raw_xls) - 1
        print(f"\033[{len(_xlsx_candidates) + 3}A\033[J", end="")
        print("\nMeerdere Excel-bestanden gevonden met de juiste kolomkoppen:")
        for _i, (_, _fn_c) in enumerate(_xlsx_candidates, 1):
            if _i - 1 == _xlsx_idx:
                print(f"  {_HL}{_i}) [{_fn_c}]{_RS}")
            else:
                print(f"  {_i}) [{_fn_c}]")
        print()
        print(f"Welk bestand wil je gebruiken? (1-{len(_xlsx_candidates)}): {_raw_xls}")
        _xlsx_info: "tuple[Path, str] | None" = _xlsx_candidates[_xlsx_idx]
    else:
        _xlsx_info = _xlsx_candidates[0]
elif _xlsx_candidates:
    _xlsx_info = _xlsx_candidates[0]
else:
    _xlsx_info = None

try:
    while True:
        semester_scope = choose_semester_scope()
        token = get_bearer_token()
        _user = _get_own_user_name(token)
        if _user:
            print(f"Ingelogd als: {_user}")

        # ── Stap 1: hoeveel studenten? ────────────────────────────────────────────────
        _HL = "\033[48;2;100;160;220m\033[1m"   # lichtblauw + bold
        _RS = "\033[0m"
        _pre_qty = ({"1": "1", "meer": "2"}.get(ARGS.aantal) if ARGS.aantal else None)
        _admin = ARGS.admin
        _valid_choices = {"1", "2"}
        if _admin:
            _valid_choices.update({"5", "6", "7"})
            if _xlsx_info:
                _valid_choices.update({"3", "4", "8"})

        _fn = f"[{_xlsx_info[1]}]" if _xlsx_info else ""
        _opt1 = f"  {_HL}1) Één student{_RS}" if _pre_qty == "1" else "  1) Één student"
        _opt2 = f"  {_HL}2) Meerdere studenten{_RS}" if _pre_qty == "2" else "  2) Meerdere studenten"
        _opt5_txt = "5) Toon alle studenten waarvan ik het portfolio kan zien en voeg er eentje toe aan .env"
        _opt6_txt = "6) Toon alle studenten die evaluatieverzoeken naar mij stuurden en voeg er eentje toe aan .env"
        _opt7_txt = "7) Toon studenten die evaluatieverzoeken stuurden maar niet mijn coach-student zijn, en niet in PORTFLOW_GILDE_STUDENTS_JSON staan, en voeg er eentje toe aan .env"
        _opt8_txt = lambda fn: f"8) Vul semester, tribe en gilde aan voor bestaande studenten in alle drie groepen vanuit {fn}"
        print("\nVoor hoeveel studenten wil je de evaluaties zien? (commandline: --aantal 1  of  --aantal meer)")
        print(_opt1)
        print(_opt2)
        _menu_lines = 2
        if _admin:
            if _xlsx_info:
                print(f"  3) Lees {_fn}, kies een coach, en voeg studenten toe aan .env (PORTFLOW_COACH_STUDENTS_JSON)")
                print(f"  4) Lees {_fn}, kies een gilde en voeg studenten toe aan PORTFLOW_GILDE_STUDENTS_JSON")
                _menu_lines += 2
            print(f"  {_opt5_txt}")
            print(f"  {_opt6_txt}")
            print(f"  {_opt7_txt}")
            _menu_lines += 3
            if _xlsx_info:
                print(f"  {_opt8_txt(_fn)}")
                _menu_lines += 1

        if _pre_qty:
            qty = _pre_qty
        else:
            qty = input("Keuze (of 'q' om te stoppen): ").strip()
            if qty in _valid_choices:
                # Wis inputregel + opties + vraagregel
                print(f"\033[{_menu_lines + 2}A\033[J", end="")
                _opt1h = f"  {_HL}1) Één student{_RS}" if qty == "1" else "  1) Één student"
                _opt2h = f"  {_HL}2) Meerdere studenten{_RS}" if qty == "2" else "  2) Meerdere studenten"
                print("Voor hoeveel studenten wil je de evaluaties zien? (commandline: --aantal 1  of  --aantal meer)")
                print(_opt1h)
                print(_opt2h)
                if _admin:
                    if _xlsx_info:
                        _opt3_hl = f"  {_HL}3) Lees {_fn} in{_RS}" if qty == "3" else f"  3) Lees {_fn}, kies een coach, en voeg studenten toe aan .env (PORTFLOW_COACH_STUDENTS_JSON)"
                        _opt4_hl = f"  {_HL}4) Lees {_fn}, kies een gilde ...{_RS}" if qty == "4" else f"  4) Lees {_fn}, kies een gilde en voeg studenten toe aan .env (PORTFLOW_GILDE_STUDENTS_JSON)"
                        print(_opt3_hl)
                        print(_opt4_hl)
                    _opt5h = f"  {_HL}{_opt5_txt}{_RS}" if qty == "5" else f"  {_opt5_txt}"
                    _opt6h = f"  {_HL}{_opt6_txt}{_RS}" if qty == "6" else f"  {_opt6_txt}"
                    _opt7h = f"  {_HL}{_opt7_txt}{_RS}" if qty == "7" else f"  {_opt7_txt}"
                    print(_opt5h)
                    print(_opt6h)
                    print(_opt7h)
                    if _xlsx_info:
                        _opt8h = f"  {_HL}{_opt8_txt(_fn)}{_RS}" if qty == "8" else f"  {_opt8_txt(_fn)}"
                        print(_opt8h)

        if qty == "3":
            if _xlsx_info:
                _import_students_from_excel(_xlsx_info[0])
            continue

        if qty == "4":
            if _xlsx_info:
                _import_gilde_from_excel(_xlsx_info[0])
            continue

        if qty == "5":
            _, token = _add_shared_student_to_env(token)
            continue

        if qty == "6":
            _, token = _add_invited_student_to_env(token)
            continue

        if qty == "7":
            _, token = _add_non_coach_invited_student_to_env(token)
            continue

        if qty == "8":
            if _xlsx_info:
                _enrich_env_from_excel(_xlsx_info[0])
            continue

        if qty.lower() == "q":
            print("Exiting gracefully. Goodbye!")
            exit()
        if qty not in ("1", "2"):
            print("Ongeldige keuze, probeer opnieuw.")
            continue

        # ── Stap 2a: één student ─────────────────────────────────────────────────────
        if qty == "1":
            options: list[tuple[str, str]] = [
                ("shared", "Kies een student uit een lijst van alle studenten die hun portfolio met mij hebben gedeeld"),
            ]
            if COACH_STUDENT_NAMES - {SEPARATOR_SENTINEL}:
                options.append(("coach", "Kies een student uit een lijst van al mijn coach studenten (uit .env)"))
            if TRIBE_STUDENT_NAMES - {SEPARATOR_SENTINEL}:
                options.append(("tribe", "Kies een student uit een lijst van studenten in mijn tribe (uit .env)"))
            if GILDE_STUDENT_NAMES - {SEPARATOR_SENTINEL}:
                options.append(("gilde", "Kies een student uit een lijst van studenten in mijn gilde (uit .env)"))
            options.append(("naam", "Voer de voornaam in van de gezochte student"))

            _lijst_source_map = {"alles": "shared", "coach": "coach", "tribe": "tribe", "gilde": "gilde"}
            _pre_source = _lijst_source_map.get(ARGS.lijst) if ARGS.lijst else None
            _pre_sub = next((str(i) for i, (k, _) in enumerate(options, 1) if k == _pre_source), None)

            def _print_lijst_menu(highlight_idx):
                print("\nWelke student?")
                for i, (_, label) in enumerate(options, start=1):
                    if str(i) == highlight_idx:
                        print(f"  {_HL}{i}) {label}{_RS}")
                    else:
                        print(f"  {i}) {label}")

            _print_lijst_menu(_pre_sub)

            if _pre_sub:
                sub = _pre_sub
            else:
                sub = input("Keuze: ").strip()
                if sub.isdigit() and 1 <= int(sub) <= len(options):
                    print(f"\033[{len(options) + 2}A\033[J", end="")
                    _print_lijst_menu(sub)

            if not sub.isdigit() or not (1 <= int(sub) <= len(options)):
                print("Ongeldige keuze.")
                continue
            source = options[int(sub) - 1][0]

            if source == "naam":
                voornaam = input("Voer de voornaam in: ").strip().lower()
                if not voornaam:
                    print("Geen naam ingevoerd.")
                    continue
                all_students, token = _fetch_shared(token)
                students = {n: d for n, d in all_students.items() if n.strip().split()[0].lower() == voornaam}
                if not students:
                    students = {n: d for n, d in all_students.items() if voornaam in n.lower()}
                env_order = None
                start_dates: dict = {}
            elif source == "shared":
                students, token = _fetch_shared(token)
                env_order = None
                start_dates = {}
            elif source == "coach":
                all_students, token = _fetch_shared(token)
                students = {n: d for n, d in all_students.items() if n in COACH_STUDENT_NAMES}
                env_order = [name for name, *_ in CURRENT_COACH_STUDENTS if name != SEPARATOR_SENTINEL]
                start_dates = COACH_STUDENT_START_DATES
            elif source == "tribe":
                all_students, token = _fetch_shared(token)
                students = {n: d for n, d in all_students.items() if n in TRIBE_STUDENT_NAMES}
                env_order = [name for name, *_ in CURRENT_TRIBE_STUDENTS if name != SEPARATOR_SENTINEL]
                start_dates = TRIBE_STUDENT_START_DATES
            else:  # gilde
                all_students, token = _fetch_shared(token)
                students = {n: d for n, d in all_students.items() if n in GILDE_STUDENT_NAMES}
                env_order = [name for name, *_ in CURRENT_GILDE_STUDENTS if name != SEPARATOR_SENTINEL]
                start_dates = GILDE_STUDENT_START_DATES

            if not students:
                print("Geen studenten gevonden.")
                continue

            ordered_names, number_width = student_order_and_width(students, preferred_order=env_order)

            # Bepaal tribe per student (vanuit de gekozen bronlijst)
            _tribe_src = (
                COACH_STUDENT_TRIBE if source == "coach"
                else TRIBE_STUDENT_TRIBE if source == "tribe"
                else GILDE_STUDENT_TRIBE if source == "gilde"
                else {}
            )

            # Groepeer per tribe (volgorde van eerste optreden bewaren)
            _tribe_groups: dict[str, list[tuple[int, str]]] = {}
            _global_idx = 1
            for sname in ordered_names:
                tribe_key = _tribe_src.get(sname, "")
                _tribe_groups.setdefault(tribe_key, []).append((_global_idx, sname))
                _global_idx += 1

            tribes = list(_tribe_groups.keys())
            _all_same_tribe = len(tribes) == 1

            if _all_same_tribe:
                # Gewone meerkolomsindeling
                _col_w = max(len(f"  {idx:0{number_width}d}) {n}") for idx, n in enumerate(ordered_names, 1)) + 2
                _cols = max(1, 80 // _col_w)
            else:
                # Per-tribe kolombreedte (afgestemd op de langste naam binnen die tribe)
                _tribe_col_w: dict[str, int] = {}
                for t in tribes:
                    grp = _tribe_groups[t]
                    _tribe_col_w[t] = max(
                        max(len(f"{idx:0{number_width}d}) {n}") for idx, n in grp),
                        len(t),
                    ) + 2  # 2 padding

                # Greedy batching: voeg tribes toe aan huidige batch zolang totale breedte ≤ 80
                # (2 spaties inspringing + kolombreedte per tribe)
                tribe_batches: list[list[str]] = []
                current_batch: list[str] = []
                current_width = 2  # de vaste "  " inspringing
                for t in tribes:
                    tw = _tribe_col_w[t]
                    if current_batch and current_width + tw > 80:
                        tribe_batches.append(current_batch)
                        current_batch = [t]
                        current_width = 2 + tw
                    else:
                        current_batch.append(t)
                        current_width += tw
                if current_batch:
                    tribe_batches.append(current_batch)

            def _print_student_list(highlight_name=None):
                print("\nStudenten:")
                if _all_same_tribe:
                    for i, (idx, sname) in enumerate(list(_tribe_groups.values())[0], start=0):
                        display_sname = anonymize_name(sname) if ARGS.anoniem else sname
                        entry_text = f"{idx:0{number_width}d}) {display_sname}"
                        entry = f"  {entry_text}"
                        padded = f"{entry:<{_col_w}}"
                        display = f"  {_HL}{entry_text}{_RS}{' ' * (_col_w - len(entry))}" if highlight_name and sname == highlight_name else padded
                        end = "\n" if (i + 1) % _cols == 0 or idx == len(ordered_names) else ""
                        print(display, end=end)
                    if len(ordered_names) % _cols != 0:
                        print()
                else:
                    for batch in tribe_batches:
                        # Header
                        header = "  " + "".join(f"{t:<{_tribe_col_w[t]}}" for t in batch)
                        print(header.rstrip())
                        print(("  " + "".join(f"{'-' * len(t):<{_tribe_col_w[t]}}" for t in batch)).rstrip())
                        # Rijen
                        max_rows = max(len(_tribe_groups[t]) for t in batch)
                        for row in range(max_rows):
                            line = "  "
                            for t in batch:
                                grp = _tribe_groups[t]
                                if row < len(grp):
                                    idx, sname = grp[row]
                                    display_sname = anonymize_name(sname) if ARGS.anoniem else sname
                                    cell = f"{idx:0{number_width}d}) {display_sname}"
                                    padded = f"{cell:<{_tribe_col_w[t]}}"
                                    cell_out = f"{_HL}{cell}{_RS}{' ' * (_tribe_col_w[t] - len(cell))}" if highlight_name and sname == highlight_name else padded
                                else:
                                    cell_out = " " * _tribe_col_w[t]
                                line += cell_out
                            print(line.rstrip())
                        print()

            if _all_same_tribe:
                _list_lines = 2 + (len(ordered_names) + _cols - 1) // _cols + (1 if len(ordered_names) % _cols != 0 else 0)
            else:
                _list_lines = 2 + sum(max(len(_tribe_groups[t]) for t in batch) + 3 for batch in tribe_batches)

            _pre_student = ARGS.student
            if not _pre_student:
                _print_student_list()

            selection = _pre_student or input("Kies een student (nummer of naam): ").strip()
            name = resolve_student_selection(selection, ordered_names, students)
            if name and selection.isdigit():
                if not _pre_student:
                    print(f"\033[{_list_lines + 1}A\033[J", end="")
                _print_student_list(highlight_name=name)
                print(f"Kies een student (nummer of naam): {selection}")
            if not name:
                print("Student niet gevonden.")
                continue

            # Tabel met één rij
            override = start_dates.get(name)
            _single_results = collect_results(token, name, students[name], semester_scope, override_start=override, include_ungraded=bool(ARGS.vraagtekens))
            if _single_results == "TOKEN_EXPIRED":
                print("Token verlopen, terug naar hoofdmenu.")
                continue
            _sem_map_s = COACH_STUDENT_SEMESTER if source == "coach" else (TRIBE_STUDENT_SEMESTER if source == "tribe" else (GILDE_STUDENT_SEMESTER if source == "gilde" else {}))
            _tribe_map_s = COACH_STUDENT_TRIBE if source == "coach" else (TRIBE_STUDENT_TRIBE if source == "tribe" else (GILDE_STUDENT_TRIBE if source == "gilde" else {}))
            _gilde_map_s = COACH_STUDENT_GILDE if source == "coach" else (TRIBE_STUDENT_GILDE if source == "tribe" else (GILDE_STUDENT_GILDE if source == "gilde" else {}))
            print_coach_table(_single_results, all_names=[name], sem_map=_sem_map_s, tribe_map=_tribe_map_s, gilde_map=_gilde_map_s)

            # Gedetailleerde tekstoutput onder de tabel
            status = print_student_evaluations(token, name, students[name], semester_scope, override_start=override)
            if status == "TOKEN_EXPIRED":
                print("Token verlopen, terug naar hoofdmenu.")
                continue
            maybe_write_schema_report()
            maybe_write_pending_debug_report()

        # ── Stap 2b: meerdere studenten ────────────────────────────────────────────
        else:
            options = [
                ("shared", "Toon een tabel met alle studenten die hun portfolio met mij hebben gedeeld"),
            ]
            if COACH_STUDENT_NAMES - {SEPARATOR_SENTINEL}:
                options.append(("coach", "Toon een tabel met al mijn coach studenten (uit .env)"))
            if TRIBE_STUDENT_NAMES - {SEPARATOR_SENTINEL}:
                options.append(("tribe", "Toon een tabel met alle studenten in mijn tribe (uit .env)"))
            if GILDE_STUDENT_NAMES - {SEPARATOR_SENTINEL}:
                options.append(("gilde", "Toon een tabel met alle studenten in mijn gilde (uit .env)"))

            _lijst_source_map2 = {"alles": "shared", "coach": "coach", "tribe": "tribe", "gilde": "gilde"}
            _pre_source2 = _lijst_source_map2.get(ARGS.lijst) if ARGS.lijst else None
            _pre_sub2 = next((str(i) for i, (k, _) in enumerate(options, 1) if k == _pre_source2), None)

            def _print_groep_menu(highlight_idx):
                print("\nWelke groep studenten?")
                for i, (_, label) in enumerate(options, start=1):
                    if str(i) == highlight_idx:
                        print(f"  {_HL}{i}) {label}{_RS}")
                    else:
                        print(f"  {i}) {label}")

            _print_groep_menu(_pre_sub2)

            if _pre_sub2:
                sub = _pre_sub2
            else:
                sub = input("Keuze: ").strip()
                if sub.isdigit() and 1 <= int(sub) <= len(options):
                    print(f"\033[{len(options) + 2}A\033[J", end="")
                    _print_groep_menu(sub)

            if not sub.isdigit() or not (1 <= int(sub) <= len(options)):
                print("Ongeldige keuze.")
                continue
            source = options[int(sub) - 1][0]

            all_students, token = _fetch_shared(token)

            if source == "shared":
                names_list = sorted(all_students.keys())
                students = all_students
                inaccessible: set = set()
                all_results = []
                for idx, name in enumerate(names_list, start=1):
                    _show_progress(idx, len(names_list), name)
                    res = collect_results(token, name, students[name], semester_scope, include_ungraded=bool(ARGS.vraagtekens), inaccessible_names=inaccessible)
                    if res == "TOKEN_EXPIRED":
                        print()
                        print("Token verlopen, terug naar hoofdmenu.")
                        break
                    all_results.extend(res)
                else:
                    print()
                    print_coach_table(all_results, all_names=names_list, inaccessible_names=inaccessible)
                    maybe_write_schema_report()
                    maybe_write_pending_debug_report()

            elif source == "coach":
                students = {n: d for n, d in all_students.items() if n in COACH_STUDENT_NAMES}
                coach_names_list = [name for name, *_ in CURRENT_COACH_STUDENTS]
                real_names = [n for n in coach_names_list if n != SEPARATOR_SENTINEL]
                all_results = []
                for idx, name in enumerate(real_names, start=1):
                    if name not in students:
                        continue
                    _show_progress(idx, len(real_names), name)
                    res = collect_results(token, name, students[name], semester_scope, override_start=COACH_STUDENT_START_DATES.get(name), include_ungraded=bool(ARGS.vraagtekens))
                    if res == "TOKEN_EXPIRED":
                        print()
                        print("Token verlopen, terug naar hoofdmenu.")
                        break
                    all_results.extend(res)
                else:
                    print()
                    print_coach_table(all_results, all_names=coach_names_list, sem_map=COACH_STUDENT_SEMESTER, tribe_map=COACH_STUDENT_TRIBE, gilde_map=COACH_STUDENT_GILDE)
                    maybe_write_schema_report()
                    maybe_write_pending_debug_report()

            elif source == "tribe":
                students = {n: d for n, d in all_students.items() if n in TRIBE_STUDENT_NAMES}
                tribe_names_list = [name for name, *_ in CURRENT_TRIBE_STUDENTS]
                real_tribe_names = [n for n in tribe_names_list if n != SEPARATOR_SENTINEL]
                inaccessible = set()
                all_results = []
                for idx, name in enumerate(real_tribe_names, start=1):
                    _show_progress(idx, len(real_tribe_names), name)
                    if name not in students:
                        inaccessible.add(name)
                        continue
                    res = collect_results(token, name, students[name], semester_scope, override_start=TRIBE_STUDENT_START_DATES.get(name), include_ungraded=bool(ARGS.vraagtekens), inaccessible_names=inaccessible)
                    if res == "TOKEN_EXPIRED":
                        print()
                        print("Token verlopen, terug naar hoofdmenu.")
                        break
                    all_results.extend(res)
                else:
                    print()
                    print_coach_table(all_results, all_names=tribe_names_list, inaccessible_names=inaccessible, sem_map=TRIBE_STUDENT_SEMESTER, tribe_map=TRIBE_STUDENT_TRIBE, gilde_map=TRIBE_STUDENT_GILDE, coach_map=TRIBE_STUDENT_COACH)
                    maybe_write_schema_report()
                    maybe_write_pending_debug_report()

            else:  # gilde
                students = {n: d for n, d in all_students.items() if n in GILDE_STUDENT_NAMES}
                gilde_names_list = [name for name, *_ in CURRENT_GILDE_STUDENTS]
                real_gilde_names = [n for n in gilde_names_list if n != SEPARATOR_SENTINEL]
                inaccessible = set()
                all_results = []
                for idx, name in enumerate(real_gilde_names, start=1):
                    _show_progress(idx, len(real_gilde_names), name)
                    if name not in students:
                        inaccessible.add(name)
                        continue
                    res = collect_results(token, name, students[name], semester_scope, override_start=GILDE_STUDENT_START_DATES.get(name), include_ungraded=bool(ARGS.vraagtekens), inaccessible_names=inaccessible)
                    if res == "TOKEN_EXPIRED":
                        print()
                        print("Token verlopen, terug naar hoofdmenu.")
                        break
                    all_results.extend(res)
                else:
                    print()
                    print_coach_table(all_results, all_names=gilde_names_list, inaccessible_names=inaccessible, sem_map=GILDE_STUDENT_SEMESTER, tribe_map=GILDE_STUDENT_TRIBE, gilde_map=GILDE_STUDENT_GILDE, coach_map=GILDE_STUDENT_COACH)
                    maybe_write_schema_report()
                    maybe_write_pending_debug_report()

        break

except KeyboardInterrupt:
    print("\nKeyboard interrupt detected. Exiting gracefully. Goodbye!")
    exit()
